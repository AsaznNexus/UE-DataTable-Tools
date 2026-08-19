# -*- coding: utf-8 -*-
"""
导出并拆表 v2.4（导出集成_v11，UI 选表版，UE 5.7 兼容）
流程：
  1. 扫描 UE 全部 DataTable，并按目录分组
  2. 在 UI 中搜索、勾选本次要导出的表
  3. 导出勾选表            → DataTables_Export
  4. 镜像备份              → DataTables_Export_留档
  5. 拆分勾选表            → DataTables_Cehua
  6. 根据 UE 扫描结果生成全量表目录.txt
"""

import unreal
import os
import re
import csv as _csv
import json
import shutil
import datetime
import sys

# ── 依赖检测：tkinter ─────────────────────────────────
try:
    import tkinter as tk
except ImportError:
    unreal.log_error("未检测到 tkinter")
    unreal.EditorDialog.show_message(
        "缺少依赖：tkinter",
        "当前 UE 内置 Python 环境未包含 tkinter 模块，\n"
        "无法弹出导出选择界面，请联系技术美术检查 Python 环境配置。",
        unreal.AppMsgType.OK
    )
    raise SystemExit("缺少依赖 tkinter")

# ── 依赖检测 ──────────────────────────────────────────
def _ensure_openpyxl():
    try:
        import openpyxl
        unreal.log(f"openpyxl 已就绪 (版本 {openpyxl.__version__})")
        return True
    except ImportError:
        unreal.log_error("未检测到 openpyxl")
        unreal.EditorDialog.show_message(
            "缺少依赖：openpyxl",
            "请先运行桌面上的「安装环境.py」脚本安装依赖，\n"
            "安装完成后再运行本脚本。",
            unreal.AppMsgType.OK
        )
        return False

if not _ensure_openpyxl():
    raise SystemExit("缺少依赖 openpyxl，脚本终止")

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles.borders import Border, Side

# ── 路径配置 ──────────────────────────────────────────
BASE_DIR     = os.path.dirname(os.path.abspath(__file__)).replace("\\", "/")
EXPORT_BASE  = f"{BASE_DIR}/DataTables_Export/"
ARCHIVE_BASE = f"{BASE_DIR}/DataTables_Export_留档/"
CEHUA_BASE   = f"{BASE_DIR}/DataTables_Cehua/"
CATALOG_FILE = f"{BASE_DIR}/全量表目录.txt"
SELECTION_CONFIG_FILE = f"{BASE_DIR}/DataTable导出选择.json"
UE_BASE_PATH = "/Game/GDataTables"
TYPE_SHEET_NAME = "__UE_META__"

# ── Source Control 检出状态缓存 ─────────────────────────
# key: DataTable相对路径, value: 状态信息
SOURCE_CONTROL_STATUS = {}
STRUCT_SPLIT_FIELDS = {
    ("AnalyticsActionDefineT", "ActionData"),
    ("InGameScoreGroupDefineT", "ScoreColor"),
    ("InGameScoreActionDefineT", "Rule"),
}
# 只允许这些表在原始 CSV 尚未经过 Excel 单元格时直接拆分。
# 用精确表名控制范围，避免改变其他表的导出链路。
DIRECT_CSV_SPLIT_TABLES = {"DifficultyDefineT"}

# DefenseRewardDefineT_S 使用专用二级拆分规则：
# ScoreRewardPool -> FixedBoxList / RandomDropBoxNum / BoxList
# 其中每个 GiftBoxTid、Weight 和数量值都单独占用一个单元格。
DEFENSE_REWARD_TABLE = "DefenseRewardDefineT_S"
DEFENSE_REWARD_FIELD = "ScoreRewardPool"
DEFENSE_REWARD_MARKER = "DEFENSE_REWARD"

# ── 样式配置 ──────────────────────────────────────────
STYLE_ROW1 = ("2F5496", "FFFFFF")  # 深蓝底 白字
STYLE_ROW2 = ("4472C4", "FFFFFF")  # 中蓝底 白字
STYLE_ROW3 = ("D9E1F2", "000000")  # 浅蓝底 黑字（字段类型）
THICK_SIDE  = Side(style="medium")
NO_SIDE     = Side(style=None)


# ══════════════════════════════════════════════════════
# 导出表扫描 + UI 选择
# ══════════════════════════════════════════════════════


def get_source_control_status(asset_path):
    """
    获取UE Source Control状态。
    返回:
      {
        "status": "正常/本地修改/已检出/锁定",
        "owner": "用户名"
      }
    """
    result = {
        "status": "未知",
        "owner": ""
    }

    try:
        if not unreal.SourceControlHelpers.is_enabled():
            result["status"] = "未启用源码管理"
            return result

        state = unreal.SourceControlHelpers.get_state(asset_path, True)
        if state is None:
            return result

        # 不同UE版本字段存在差异，全部使用安全访问
        if state.is_checked_out():
            result["status"] = "已检出"
            try:
                result["owner"] = str(state.get_checked_out_user())
            except Exception:
                pass
        elif state.is_modified():
            result["status"] = "本地修改"
        elif state.is_locked():
            result["status"] = "锁定"

        return result

    except Exception as e:
        unreal.log_warning(f"读取Source Control状态失败 {asset_path}: {e}")
        return result


def scan_exportable_tables():
    """
    扫描 UE_BASE_PATH 下全部 DataTable，按顶层文件夹分组。
    返回 (grouped, all_entries)：
      grouped = {分组名: [相对资产路径, ...]}
      all_entries = 全部相对资产路径（排序后）
    """
    groups = {}
    entries = []

    unreal.log(f"正在扫描 UE DataTable：{UE_BASE_PATH}")
    asset_list = unreal.EditorAssetLibrary.list_assets(UE_BASE_PATH, recursive=True)

    for asset_path in asset_list:
        try:
            asset_data = unreal.EditorAssetLibrary.find_asset_data(asset_path)
            class_name = str(asset_data.asset_class_path.asset_name)
            if class_name != "DataTable":
                continue

            package_name = str(asset_data.package_name).replace("\\", "/")
            prefix = UE_BASE_PATH.rstrip("/") + "/"
            if not package_name.startswith(prefix):
                continue

            entry = package_name[len(prefix):].strip("/")
            if not entry:
                continue

            entries.append(entry)

            # 查询当前DataTable的源码管理状态
            SOURCE_CONTROL_STATUS[entry] = get_source_control_status(asset_path)

            group = entry.split("/", 1)[0] if "/" in entry else "（未分类）"
            groups.setdefault(group, []).append(entry)
        except Exception as e:
            unreal.log_warning(f"读取资产信息失败，已跳过：{asset_path} -> {e}")

    grouped = {
        group: sorted(group_entries, key=str.lower)
        for group, group_entries in sorted(groups.items(), key=lambda item: item[0].lower())
    }
    all_entries = sorted(set(entries), key=str.lower)
    unreal.log(f"扫描完成：找到 {len(all_entries)} 个 DataTable")
    return grouped, all_entries


def load_last_selection():
    """读取上次确认导出的表，作为 UI 默认勾选。"""
    if not os.path.exists(SELECTION_CONFIG_FILE):
        return set()
    try:
        with open(SELECTION_CONFIG_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        selected = data.get("selected_tables", []) if isinstance(data, dict) else []
        return {str(entry).replace("\\", "/") for entry in selected if entry}
    except Exception as e:
        unreal.log_warning(f"读取上次导出选择失败：{e}")
        return set()


def save_last_selection(selected):
    """保存本次确认导出的表。"""
    try:
        with open(SELECTION_CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(
                {"selected_tables": list(selected)},
                f,
                ensure_ascii=False,
                indent=2,
            )
    except Exception as e:
        unreal.log_warning(f"保存本次导出选择失败：{e}")


def select_export_tables_gui(grouped, preselected=None):
    """
    显示分组勾选窗口，支持多关键词搜索、按组选择、仅显示已选。
    返回 (selected, open_output_folder)：关闭或取消时 selected 为 None。
    """
    preselected = preselected or set()
    result = {"selected": None, "open_output_folder": True}

    root = tk.Tk()
    root.title("选择要导出的表")
    root.geometry("640x760")
    root.minsize(540, 500)
    root.attributes("-topmost", True)

    top = tk.Frame(root)
    top.pack(fill="x", padx=10, pady=(10, 5))

    tk.Label(top, text="搜索：").pack(side="left")
    search_var = tk.StringVar()
    search_entry = tk.Entry(top, textvariable=search_var)
    search_entry.pack(side="left", fill="x", expand=True, padx=(0, 8))

    show_selected_var = tk.BooleanVar(value=False)
    tk.Checkbutton(top, text="仅显示已选", variable=show_selected_var).pack(side="left", padx=(0, 8))

    count_label = tk.Label(root, text="", anchor="w")
    count_label.pack(fill="x", padx=10)

    list_frame = tk.Frame(root)
    list_frame.pack(fill="both", expand=True, padx=10, pady=5)

    canvas = tk.Canvas(list_frame, highlightthickness=0)
    scrollbar = tk.Scrollbar(list_frame, orient="vertical", command=canvas.yview)
    inner = tk.Frame(canvas)
    inner_window = canvas.create_window((0, 0), window=inner, anchor="nw")

    inner.bind("<Configure>", lambda _e: canvas.configure(scrollregion=canvas.bbox("all")))
    canvas.bind("<Configure>", lambda e: canvas.itemconfigure(inner_window, width=e.width))
    canvas.configure(yscrollcommand=scrollbar.set)
    canvas.pack(side="left", fill="both", expand=True)
    scrollbar.pack(side="right", fill="y")

    def _on_mousewheel(event):
        canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    canvas.bind_all("<MouseWheel>", _on_mousewheel)

    vars_by_entry = {}
    rows_by_entry = {}
    group_headers = {}
    batch_state = {"active": False}

    def get_search_tokens():
        return search_var.get().strip().lower().split()

    def entry_matches(entry):
        tokens = get_search_tokens()
        search_text = entry.lower()
        if not all(token in search_text for token in tokens):
            return False
        if show_selected_var.get() and not vars_by_entry[entry].get():
            return False
        return True

    def get_visible_entries():
        return [entry for entry in vars_by_entry if entry_matches(entry)]

    def update_count():
        visible = get_visible_entries()
        visible_selected = sum(bool(vars_by_entry[e].get()) for e in visible)
        total_selected = sum(bool(v.get()) for v in vars_by_entry.values())
        count_label.config(
            text=(
                f"找到 {len(visible)} 个，当前结果已选 {visible_selected} 个；"
                f"全部 {len(vars_by_entry)} 个，已选 {total_selected} 个"
            )
        )

    def apply_filter(*_):
        for group, entries in grouped.items():
            group_headers[group].pack_forget()
            for entry in entries:
                rows_by_entry[entry].pack_forget()

        for group, entries in grouped.items():
            visible = [entry for entry in entries if entry_matches(entry)]
            if not visible:
                continue
            group_headers[group].pack(fill="x", pady=(8, 2))
            for entry in visible:
                rows_by_entry[entry].pack(fill="x", anchor="w", padx=(24, 0))

        root.update_idletasks()
        canvas.configure(scrollregion=canvas.bbox("all"))
        canvas.yview_moveto(0)
        update_count()

    def on_selection_changed(*_):
        if batch_state["active"]:
            return
        if show_selected_var.get():
            apply_filter()
        else:
            update_count()

    def set_entries(entries, value):
        batch_state["active"] = True
        try:
            for entry in entries:
                vars_by_entry[entry].set(value)
        finally:
            batch_state["active"] = False
        if show_selected_var.get():
            apply_filter()
        else:
            update_count()

    for group, entries in grouped.items():
        header = tk.Frame(inner, bg="#2F5496")
        header.pack(fill="x", pady=(8, 2))
        tk.Label(
            header,
            text=f"—— {group} ——",
            bg="#2F5496",
            fg="white",
            anchor="w",
        ).pack(side="left", fill="x", expand=True, padx=6, pady=3)
        tk.Button(
            header,
            text="组清空",
            command=lambda es=entries: set_entries(es, False),
        ).pack(side="right", padx=(2, 6), pady=2)
        tk.Button(
            header,
            text="组全选",
            command=lambda es=entries: set_entries(es, True),
        ).pack(side="right", padx=2, pady=2)
        group_headers[group] = header

        for entry in entries:
            var = tk.BooleanVar(value=(entry in preselected))
            short_name = entry.split("/", 1)[1] if "/" in entry else entry

            # UI显示Source Control状态
            sc_info = SOURCE_CONTROL_STATUS.get(entry, {})
            sc_status = sc_info.get("status", "未知")
            sc_owner = sc_info.get("owner", "")
            if sc_status == "已检出" and sc_owner:
                display_name = f"{short_name}   🔒 {sc_status}({sc_owner})"
            elif sc_status != "未知":
                display_name = f"{short_name}   [{sc_status}]"
            else:
                display_name = short_name

            cb = tk.Checkbutton(
                inner,
                text=display_name,
                variable=var,
                anchor="w",
                justify="left",
            )
            cb.pack(fill="x", anchor="w", padx=(24, 0))
            vars_by_entry[entry] = var
            rows_by_entry[entry] = cb
            var.trace_add("write", on_selection_changed)

    search_var.trace_add("write", apply_filter)
    show_selected_var.trace_add("write", apply_filter)

    def select_all_visible():
        set_entries(get_visible_entries(), True)

    def select_none_visible():
        set_entries(get_visible_entries(), False)

    tk.Button(top, text="全选", command=select_all_visible).pack(side="left", padx=2)
    tk.Button(top, text="全不选", command=select_none_visible).pack(side="left", padx=2)

    options = tk.Frame(root)
    options.pack(fill="x", padx=10, pady=(4, 0))
    open_folder_var = tk.BooleanVar(value=True)
    tk.Checkbutton(
        options,
        text="完成后打开 DataTables_Cehua 文件夹",
        variable=open_folder_var,
        anchor="w",
    ).pack(side="left")

    bottom = tk.Frame(root)
    bottom.pack(fill="x", padx=10, pady=10)

    def on_confirm():
        selected = [entry for entry, var in vars_by_entry.items() if var.get()]
        result["selected"] = selected
        result["open_output_folder"] = bool(open_folder_var.get())
        root.destroy()

    def on_cancel():
        result["selected"] = None
        root.destroy()

    tk.Button(bottom, text="取消", width=10, command=on_cancel).pack(side="right", padx=5)
    tk.Button(
        bottom,
        text="确定导出",
        width=10,
        command=on_confirm,
        bg="#2F5496",
        fg="white",
    ).pack(side="right", padx=5)

    root.protocol("WM_DELETE_WINDOW", on_cancel)
    update_count()
    search_entry.focus_set()
    root.mainloop()

    return result["selected"], result["open_output_folder"]


# ══════════════════════════════════════════════════════
# UE CSV 格式处理工具
# ══════════════════════════════════════════════════════

def _normalize_ue_csv_cell(s):
    """
    把 UE CSV 单元格里的值标准化：
    - UE CSV 数组格式 ((k=v,...),(k=v,...)) 原样保留，
      拆表逻辑的 parse_ue_struct / collect_subkeys 能直接识别
    - 非数组值原样返回
    """
    if not s:
        return s
    s = s.strip()
    return s


def _to_snake_case(name):
    """把 UE 属性名转换为 Python Editor Property 常用的 snake_case。"""
    s1 = re.sub(r'(.)([A-Z][a-z]+)', r'\1_\2', str(name))
    return re.sub(r'([a-z0-9])([A-Z])', r'\1_\2', s1).lower()


def _classify_container_value(value):
    """根据 RowStruct 默认实例中的实际值判断容器类型。"""
    type_name = type(value).__name__.lower()
    if type_name in ("array", "set"):
        return "ARRAY"
    if type_name == "map":
        return "MAP"

    try:
        if isinstance(value, (unreal.Array, unreal.Set)):
            return "ARRAY"
        if isinstance(value, unreal.Map):
            return "MAP"
    except Exception:
        pass
    return None


def get_row_struct_field_types(data_table, export_headers):
    """
    从 DataTable 的 RowStruct 默认实例读取字段真实类型。
    返回 {CSV导出字段名: "ARRAY" | "MAP"}；无法反射的字段留给原有内容分析兜底。
    """
    result = {}
    try:
        row_struct = unreal.DataTableFunctionLibrary.get_data_table_row_struct(data_table)
        if row_struct is None:
            return result

        struct_name = str(row_struct.get_name())
        candidates = [struct_name]
        if struct_name.startswith("F") and len(struct_name) > 1:
            candidates.append(struct_name[1:])

        struct_type = None
        for candidate in candidates:
            struct_type = getattr(unreal, candidate, None)
            if struct_type is not None:
                break
        if struct_type is None:
            unreal.log_warning(f"RowStruct 未生成 Python 类型，改用内容识别: {struct_name}")
            return result

        default_row = struct_type()
        for export_name in export_headers:
            raw_name = unreal.DataTableFunctionLibrary.get_data_table_column_name_from_export_name(
                data_table, str(export_name)
            )
            raw_name = str(raw_name) if raw_name is not None else str(export_name)
            value = None
            found = False
            for prop_name in (raw_name, _to_snake_case(raw_name)):
                try:
                    value = default_row.get_editor_property(prop_name)
                    found = True
                    break
                except Exception:
                    continue
            if found:
                marker = _classify_container_value(value)
                if marker:
                    result[str(export_name)] = marker
    except Exception as e:
        unreal.log_warning(f"读取 RowStruct 字段类型失败，改用内容识别: {e}")
    return result


# ══════════════════════════════════════════════════════
# 【第一部分】导出：UE → DataTables_Export
# ══════════════════════════════════════════════════════

def run_export(table_list):
    # 清空并重建导出文件夹
    if os.path.exists(EXPORT_BASE):
        shutil.rmtree(EXPORT_BASE)
        unreal.log("已清空导出文件夹")
    os.makedirs(EXPORT_BASE, exist_ok=True)

    # 清空策划文件夹
    if os.path.exists(CEHUA_BASE):
        shutil.rmtree(CEHUA_BASE)
        unreal.log("已清空策划文件夹")
    os.makedirs(CEHUA_BASE, exist_ok=True)

    total         = len(table_list)
    success_count = 0
    fail_list     = []
    exported_list = []

    with unreal.ScopedSlowTask(total, "正在导出 DataTable...") as task:
        task.make_dialog(True)

        for entry in table_list:
            asset_name    = os.path.basename(entry)
            relative      = entry
            relative_dir  = os.path.dirname(entry)
            ue_asset_path = f"{UE_BASE_PATH}/{entry}"

            task.enter_progress_frame(1, f"导出中 ({success_count}/{total})：{asset_name}")

            if task.should_cancel():
                unreal.log_warning("用户取消了导出操作")
                break

            output_dir = os.path.join(EXPORT_BASE, relative_dir)
            os.makedirs(output_dir, exist_ok=True)

            try:
                data_table = unreal.load_asset(ue_asset_path)
                if data_table is None:
                    fail_list.append(relative)
                    unreal.log_error(f"找不到UE资产: {ue_asset_path}")
                    continue

                # UE 5.7 兼容：使用 CSV 接口，绕开 JSON 序列化崩溃
                temp_csv = os.path.join(output_dir, f"{asset_name}_temp.csv")
                csv_ok = unreal.DataTableFunctionLibrary.export_data_table_to_csv_file(data_table, temp_csv)
                if not csv_ok or not os.path.exists(temp_csv):
                    fail_list.append(relative)
                    unreal.log_error(f"CSV导出失败: {relative}")
                    continue

                with open(temp_csv, "r", encoding="utf-8") as f:
                    reader = _csv.reader(f)
                    all_csv_rows = list(reader)
                os.remove(temp_csv)

                if not all_csv_rows:
                    fail_list.append(relative)
                    unreal.log_error(f"CSV为空: {relative}")
                    continue

                csv_headers = all_csv_rows[0]   # 第一行：列名
                csv_data    = all_csv_rows[1:]   # 其余：数据行

                # 从 UE RowStruct 读取真实容器类型。
                field_types = get_row_struct_field_types(data_table, csv_headers[1:])
                xlsx_path = os.path.join(output_dir, f"{asset_name}.xlsx")
                if asset_name in DIRECT_CSV_SPLIT_TABLES:
                    # DifficultyDefineT 的超长数组直接从完整 CSV 内存拆列，
                    # 避开 Excel 单元格 32767 字符上限。
                    msg = process_csv_rows_directly(
                        asset_name, csv_headers, csv_data, field_types, xlsx_path
                    )
                    unreal.log(f"定向直拆: {relative} - {msg}")
                else:
                    # 其他表完全保留 v7 的普通导出流程。
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = asset_name[:31]

                    # 写入隐藏元数据表。
                    ws_meta = wb.create_sheet(TYPE_SHEET_NAME)
                    ws_meta.append(["FieldName", "FieldType"])
                    for col_name in csv_headers[1:]:
                        ws_meta.append([col_name, field_types.get(str(col_name), None)])
                    ws_meta.sheet_state = "hidden"

                    header_fill = PatternFill("solid", fgColor="2F5496")
                    header_font = Font(color="FFFFFF", bold=True)

                    # 写三行表头（第一列是 RowName，对应旧版的 "---"）
                    ws.cell(row=1, column=1, value="---").fill = header_fill
                    ws.cell(row=1, column=1).font = header_font
                    for col_idx, col_name in enumerate(csv_headers[1:], start=2):
                        cell = ws.cell(row=1, column=col_idx, value=col_name)
                        cell.fill = header_fill
                        cell.font = header_font

                    # 第二行保留为拆分子字段表头；第三行固定保存字段类型。
                    ws.cell(row=3, column=1, value="#FieldType")
                    for col_idx, col_name in enumerate(csv_headers[1:], start=2):
                        ws.cell(row=3, column=col_idx,
                                value=field_types.get(str(col_name), None))
                    style_header_row(ws, 2, *STYLE_ROW2)
                    style_header_row(ws, 3, *STYLE_ROW3)

                    # 写数据（UE CSV 第一列是 RowName，其余原样写入）
                    for row_idx, row in enumerate(csv_data, start=4):
                        for col_idx, val in enumerate(row, start=1):
                            ws.cell(row=row_idx, column=col_idx, value=val if val else None)

                    for col in ws.columns:
                        max_len = max((len(str(c.value or "")) for c in col), default=0)
                        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

                    wb.save(xlsx_path)

                success_count += 1
                exported_list.append(relative)
                unreal.log(f"导出成功: {relative} ({len(csv_data)} 行)")

            except Exception as e:
                fail_list.append(relative)
                unreal.log_error(f"导出失败: {relative} -> {e}")

    unreal.log(f"\n导出完成：成功 {success_count} 个，失败 {len(fail_list)} 个")
    return success_count, fail_list, exported_list


# ══════════════════════════════════════════════════════
# 【第二部分】留档 + 全量表目录
# ══════════════════════════════════════════════════════

def run_archive(exported_list, catalog_entries=None):
    """生成本次导出留档，并根据 UE 扫描结果生成全量表目录。"""
    catalog_entries = catalog_entries if catalog_entries is not None else exported_list
    with unreal.ScopedSlowTask(2, "正在生成留档和目录...") as task:
        task.make_dialog(False)

        task.enter_progress_frame(1, "生成留档镜像...")
        if os.path.exists(ARCHIVE_BASE):
            shutil.rmtree(ARCHIVE_BASE)
        shutil.copytree(EXPORT_BASE, ARCHIVE_BASE)
        unreal.log(f"留档镜像已生成: {ARCHIVE_BASE}")

        task.enter_progress_frame(1, "生成全量表目录...")
        dir_groups = {}
        for relative in sorted(catalog_entries):
            folder = os.path.dirname(relative) or "（根目录）"
            if folder not in dir_groups:
                dir_groups[folder] = []
            dir_groups[folder].append(relative)

        now_str = datetime.datetime.now().strftime("%Y/%m/%d %H:%M")
        lines = [
            f"# 全量表目录 - {now_str}",
            f"# 共 {len(catalog_entries)} 个表",
            f"# 导出范围请在导出集成 UI 中勾选",
            "",
        ]
        for folder in sorted(dir_groups.keys()):
            lines.append(f"# ── {folder} ──")
            for relative in dir_groups[folder]:
                lines.append(relative)
            lines.append("")

        os.makedirs(os.path.dirname(CATALOG_FILE), exist_ok=True)
        with open(CATALOG_FILE, "w", encoding="utf-8") as f:
            f.write("\n".join(lines))

        unreal.log(f"全量表目录已生成: {CATALOG_FILE} ({len(catalog_entries)} 个表)")


# ══════════════════════════════════════════════════════
# 【第三部分】拆表核心逻辑
# ══════════════════════════════════════════════════════

def parse_ue_struct(s):
    """解析 (k=v,k=v) 格式，返回 (keys列表, kv字典)（引号内的括号/逗号不参与计数）"""
    s = s.strip()
    if s.startswith("(") and s.endswith(")"):
        s = s[1:-1]
    keys, kv = [], {}
    depth, current = 0, []
    in_q = False
    for ch in s:
        if ch == '"':
            in_q = not in_q
            current.append(ch)
        elif ch in "([" and not in_q:
            depth += 1
            current.append(ch)
        elif ch in ")]" and not in_q:
            depth -= 1
            current.append(ch)
        elif ch == "," and depth == 0 and not in_q:
            token = "".join(current).strip()
            if "=" in token:
                k, _, v = token.partition("=")
                keys.append(k.strip())
                kv[k.strip()] = v.strip().strip('"')
            current = []
        else:
            current.append(ch)
    token = "".join(current).strip()
    if "=" in token:
        k, _, v = token.partition("=")
        keys.append(k.strip())
        kv[k.strip()] = v.strip().strip('"')
    return keys, kv


def _split_struct_assignments(inner):
    """按顶层逗号拆分单结构体字段，保留引号和嵌套结构原文。"""
    tokens, current = [], []
    depth = 0
    in_q = False
    escaped = False
    for ch in inner:
        if in_q:
            current.append(ch)
            if escaped:
                escaped = False
            elif ch == "\\":
                escaped = True
            elif ch == '"':
                in_q = False
        elif ch == '"':
            in_q = True
            current.append(ch)
        elif ch in "([{":
            depth += 1
            current.append(ch)
        elif ch in ")]}":
            depth -= 1
            current.append(ch)
        elif ch == "," and depth == 0:
            token = "".join(current).strip()
            if token:
                tokens.append(token)
            current = []
        else:
            current.append(ch)
    token = "".join(current).strip()
    if token:
        tokens.append(token)
    return tokens


def parse_single_ue_struct(s):
    """
    解析单个 UE 结构体，返回 [(子字段, 单元格值, 类型标记), ...]。
    类型标记用于第三行：STRUCT_STRING / STRUCT_VALUE / STRUCT_RAW。
    """
    if s is None:
        return None
    text = str(s).strip()
    if not (text.startswith("(") and text.endswith(")")):
        return None
    inner = text[1:-1].strip()
    if not inner or inner.startswith("("):
        return None

    result = []
    for token in _split_struct_assignments(inner):
        if "=" not in token:
            return None
        key, _, raw_value = token.partition("=")
        key = key.strip()
        raw_value = raw_value.strip()
        if not key:
            return None
        if len(raw_value) >= 2 and raw_value.startswith('"') and raw_value.endswith('"'):
            marker = "STRUCT_STRING"
            cell_value = raw_value[1:-1]
        elif raw_value.startswith(("(", "[", "{")):
            marker = "STRUCT_RAW"
            cell_value = raw_value
        else:
            marker = "STRUCT_VALUE"
            cell_value = raw_value
        result.append((key, cell_value, marker))
    return result if result else None


def analyze_selected_struct_columns(asset_name, headers, data_rows):
    """只分析配置中指定表、指定字段的单结构体。"""
    result = {}
    for ci, field in enumerate(headers):
        field_name = str(field) if field is not None else ""
        if (asset_name, field_name) not in STRUCT_SPLIT_FIELDS:
            continue

        subkeys = []
        modes = {}
        found = False
        for row in data_rows:
            value = row[ci] if ci < len(row) else None
            parsed = parse_single_ue_struct(value)
            if parsed is None:
                continue
            found = True
            for subkey, _, marker in parsed:
                if subkey not in subkeys:
                    subkeys.append(subkey)
                previous = modes.get(subkey)
                if previous is None:
                    modes[subkey] = marker
                elif "STRUCT_RAW" in (previous, marker):
                    modes[subkey] = "STRUCT_RAW"
                elif "STRUCT_STRING" in (previous, marker):
                    modes[subkey] = "STRUCT_STRING"
                else:
                    modes[subkey] = "STRUCT_VALUE"

        if found and subkeys:
            result[ci] = {
                "field": field_name,
                "subkeys": subkeys,
                "modes": modes,
            }
    return result


def collect_subkeys(all_items):
    """从数组元素列表里收集所有子字段名，保持顺序"""
    seen, keys = set(), []
    for item in all_items:
        if isinstance(item, str) and item.strip().startswith("("):
            ks, _ = parse_ue_struct(item)
            for k in ks:
                if k not in seen:
                    seen.add(k)
                    keys.append(k)
    return keys if keys else None


def _split_ue_array_items(inner):
    """把 (k=v,...),(k=v,...) 拆成带括号的元素列表（引号内的括号/逗号不参与计数）"""
    items, depth, current = [], 0, []
    in_q = False
    for ch in inner:
        if ch == '"':
            in_q = not in_q
            current.append(ch)
        elif ch == "(" and not in_q:
            depth += 1
            current.append(ch)
        elif ch == ")" and not in_q:
            depth -= 1
            current.append(ch)
            if depth == 0:
                token = "".join(current).strip()
                if token:
                    items.append(token)
                current = []
        elif ch == "," and depth == 0 and not in_q:
            pass  # 括号外的逗号是元素分隔符，跳过
        else:
            current.append(ch)
    return items


def _parse_kv_pair_item(item_str):
    """
    把 TMap 的 (Key, Value) 格式（括号内无等号，逗号分隔两个值）
    转换为标准 struct 字符串 (Key=key,Value=value)。
    若不符合格式则返回原字符串。
    """
    s = item_str.strip()
    if not (s.startswith("(") and s.endswith(")")):
        return item_str
    inner = s[1:-1]
    # 必须无等号，否则是普通 struct
    if "=" in inner:
        return item_str
    # 按第一个逗号分成两部分（Value 可能含逗号，取首个）
    comma_idx = inner.find(",")
    if comma_idx == -1:
        return item_str
    k = inner[:comma_idx].strip().strip('"')
    v = inner[comma_idx + 1:].strip().strip('"')
    if not k:
        return item_str
    # 重新序列化为标准 (Key=k,Value=v) 格式
    k_part = f'Key="{k}"' if not k.lstrip("-").replace(".", "", 1).isdigit() else f"Key={k}"
    v_part = f'Value="{v}"' if not v.lstrip("-").replace(".", "", 1).isdigit() else f"Value={v}"
    return f"({k_part},{v_part})"


def _try_parse_array(s):
    """
    解析单元格为列表，支持以下格式：
    1. JSON 数组 [...] — 元素可以是 dict / "(k=v)" 字符串 / "(Key,Value)" 字符串 / 标量
    2. UE CSV 结构体数组  ((k=v,...),(k=v,...))  — 双层括号
    3. UE CSV 标量数组    (v1,v2,v3)             — 单层括号，内部无等号且无嵌套括号
    所有结果统一为 "(k=v,...)" 字符串列表（结构体）或标量列表，
    供 collect_subkeys / _detect_col_kind 识别。
    """
    if not s:
        return None
    s = str(s).strip()

    # JSON 格式 [...]
    if s.startswith("[") and s.endswith("]"):
        try:
            parsed = json.loads(s)
            if isinstance(parsed, list):
                result = []
                for item in parsed:
                    if isinstance(item, str) and item.strip().startswith("("):
                        result.append(_parse_kv_pair_item(item.strip()))
                    elif isinstance(item, dict):
                        parts = []
                        for k, v in item.items():
                            if isinstance(v, str):
                                parts.append(f'{k}="{v}"')
                            else:
                                parts.append(f"{k}={v}")
                        result.append(f"({','.join(parts)})")
                    else:
                        result.append(item)
                return result
        except Exception:
            pass

    # UE CSV 格式
    if s.startswith("(") and s.endswith(")"):
        inner = s[1:-1].strip()

        if inner == "":
            # 空数组 ()，返回空列表（不拆列，analyze_array_columns 会跳过空列表）
            return None

        if inner.startswith("("):
            # 双层括号：结构体数组或 TMap
            items = _split_ue_array_items(inner)
            if items:
                return [_parse_kv_pair_item(item) for item in items]
        else:
            # 单层括号：可能是纯标量数组 (v1,v2,v3)
            # 若内部含等号则是单个 struct 值 (k=v,k=v)，不是数组，跳过
            if "=" in inner:
                return None
            items = _split_scalar_items_export(inner)
            if items and len(items) > 1:
                # 多个元素才视为数组，单元素 (v) 视为普通 struct 包装不拆
                return items  # 普通标量列表，_detect_col_kind 会判断为 tset

    return None


def _split_scalar_items_export(inner):
    """
    把 v1,v2,v3 拆成 ['v1','v2','v3']，
    正确处理带括号的嵌套（嵌套括号内的逗号不分割）。
    """
    items, depth, current = [], 0, []
    in_q = False
    for ch in inner:
        if ch == '"':
            in_q = not in_q
            current.append(ch)
        elif ch == "(" and not in_q:
            depth += 1
            current.append(ch)
        elif ch == ")" and not in_q:
            depth -= 1
            current.append(ch)
        elif ch == "," and depth == 0 and not in_q:
            token = "".join(current).strip()
            if token:
                items.append(token)
            current = []
        else:
            current.append(ch)
    token = "".join(current).strip()
    if token:
        items.append(token)
    return items


def _detect_col_kind(all_items):
    """
    检测数组列的种类，返回:
      "tmap"   — 每个元素是 (Key=...,Value=...) struct
      "tset"   — 每个元素是纯标量（数字或字符串），视为 TArray/TSet
      "struct" — 每个元素是其他 struct
      None     — 无法识别或混合格式，不拆列（原样保留）
    规则：
      - 全是 "(k=v,...)" 格式 → 判断子字段决定 tmap / struct
      - 全是非括号标量（int/float/string）→ tset
      - 混合或单元素带括号但无 = → None（单个 struct 值，不是数组）
    """
    if not all_items:
        return None

    struct_items = []
    scalar_items = []
    for item in all_items:
        if isinstance(item, str) and item.strip().startswith("("):
            struct_items.append(item)
        else:
            scalar_items.append(item)

    # 混合格式：不拆
    if struct_items and scalar_items:
        return None

    if scalar_items and not struct_items:
        # 全是非括号元素（int/float/string）→ TArray 标量，拆为 item_N
        return "tset"

    # 全是 struct 元素
    subkeys = collect_subkeys(struct_items)
    if subkeys is None:
        return None
    key_set = set(k.lower() for k in subkeys)
    if key_set == {"key", "value"}:
        return "tmap"
    return "struct"


def analyze_array_columns(headers, data_rows):
    info = {}
    for row in data_rows:
        for ci, val in enumerate(row):
            if val is None:
                continue
            parsed = _try_parse_array(val)
            if parsed is None:
                continue
            if ci not in info:
                info[ci] = {"field": headers[ci], "max_len": 0, "all_items": []}
            if len(parsed) > info[ci]["max_len"]:
                info[ci]["max_len"] = len(parsed)
            info[ci]["all_items"].extend(parsed)
    result = {}
    for ci, d in info.items():
        subkeys = collect_subkeys(d["all_items"])
        kind    = _detect_col_kind(d["all_items"])
        result[ci] = {
            "field":   d["field"],
            "max_len": d["max_len"],
            "subkeys": subkeys,
            "kind":    kind,       # "tmap" | "tset" | "struct"
        }
    return result


def build_column_plan(headers, array_cols, struct_cols):
    """
    生成列计划，plan 元素格式：
      ("normal",       ci, None,    None)
      ("tmap_key",     ci, None,    arr_idx)   — TMap Key
      ("tmap_val",     ci, None,    arr_idx)   — TMap Value
      ("tset_item",    ci, None,    arr_idx)   — TSet 标量元素
      ("array_struct", ci, subkey,  arr_idx)   — 普通结构体数组子字段
      ("array_scalar", ci, None,    arr_idx)   — 普通标量数组
      ("struct_*",     ci, subkey,  None)      — 指定单结构体子字段
    """
    row1, row2, plan = [], [], []
    for ci, field in enumerate(headers):
        if ci in struct_cols:
            sc = struct_cols[ci]
            for subkey in sc["subkeys"]:
                marker = sc["modes"].get(subkey, "STRUCT_VALUE")
                row1.append(field)
                row2.append(subkey)
                plan.append((marker.lower(), ci, subkey, None))
            continue

        if ci not in array_cols:
            row1.append(field)
            row2.append(None)
            plan.append(("normal", ci, None, None))
            continue

        ac      = array_cols[ci]
        max_len = ac["max_len"]
        subkeys = ac["subkeys"]
        kind    = ac["kind"]
        multi   = max_len > 1

        if kind == "tmap":
            # TMap：每个元素拆成 key_N / val_N 两列
            for arr_idx in range(max_len):
                row1.append(field)
                row2.append(f"key_{arr_idx}" if multi else "key")
                plan.append(("tmap_key", ci, None, arr_idx))
                row1.append(field)
                row2.append(f"val_{arr_idx}" if multi else "val")
                plan.append(("tmap_val", ci, None, arr_idx))

        elif kind == "tset":
            # TSet：每个元素一列，子字段名 item_N
            for arr_idx in range(max_len):
                row1.append(field)
                row2.append(f"item_{arr_idx}" if multi else "item")
                plan.append(("tset_item", ci, None, arr_idx))

        elif kind == "struct" and subkeys:
            # 普通结构体数组
            for arr_idx in range(max_len):
                for sk in subkeys:
                    row1.append(field)
                    row2.append(f"{sk}_{arr_idx}" if multi else sk)
                    plan.append(("array_struct", ci, sk, arr_idx))

        else:
            # 兜底：普通标量数组
            for arr_idx in range(max_len):
                row1.append(field)
                row2.append(f"value_{arr_idx}" if multi else "value")
                plan.append(("array_scalar", ci, None, arr_idx))

    return row1, row2, plan


def _get_tmap_key_value(item):
    """从 TMap 元素中取 Key / Value，兼容大小写"""
    if isinstance(item, str) and item.strip().startswith("("):
        _, kv = parse_ue_struct(item)
        # 找大小写无关的 Key / Value
        key_name = next((k for k in kv if k.lower() == "key"),   None)
        val_name = next((k for k in kv if k.lower() == "value"), None)
        return kv.get(key_name, None), kv.get(val_name, None)
    if isinstance(item, dict):
        key_name = next((k for k in item if k.lower() == "key"),   None)
        val_name = next((k for k in item if k.lower() == "value"), None)
        return item.get(key_name, None), item.get(val_name, None)
    return None, None


def expand_row(row, col_plan):
    new_row      = []
    parsed_cache = {}
    struct_cache = {}
    for ptype, ci, subkey, arr_idx in col_plan:
        val = row[ci] if ci < len(row) else None
        if ptype == "normal":
            new_row.append(val)
            continue
        if ptype in ("struct_string", "struct_value", "struct_raw"):
            if ci not in struct_cache:
                parsed = parse_single_ue_struct(val) or []
                struct_cache[ci] = {key: value for key, value, _ in parsed}
            new_row.append(struct_cache[ci].get(subkey, None))
            continue
        if ci not in parsed_cache:
            parsed_cache[ci] = _try_parse_array(val) or []
        parsed = parsed_cache[ci]
        if arr_idx >= len(parsed):
            new_row.append(None)
            continue
        item = parsed[arr_idx]

        if ptype == "tmap_key":
            k, _ = _get_tmap_key_value(item)
            new_row.append(k)

        elif ptype == "tmap_val":
            _, v = _get_tmap_key_value(item)
            new_row.append(v)

        elif ptype == "tset_item":
            # TSet 标量：直接取元素值
            if isinstance(item, str) and item.strip().startswith("("):
                # 形如 "(123)" 的包装
                inner = item.strip()[1:-1].strip()
                new_row.append(inner if inner else item)
            else:
                new_row.append(item)

        elif ptype == "array_struct":
            if isinstance(item, str) and item.strip().startswith("("):
                _, kv = parse_ue_struct(item)
                new_row.append(kv.get(subkey, None))
            elif isinstance(item, dict):
                v = item.get(subkey, None)
                new_row.append(str(v) if v is not None else None)
            else:
                new_row.append(str(item))

        elif ptype == "array_scalar":
            new_row.append(item)

    return new_row


def _defense_reward_struct_dict(value):
    """将 UE 结构体文本转为字典，嵌套数组保留原文。"""
    if isinstance(value, dict):
        return value
    if isinstance(value, str) and value.strip().startswith("("):
        _, result = parse_ue_struct(value.strip())
        return result
    return {}


def _defense_reward_struct_array(value):
    """解析 UE 结构体数组，返回字典列表。"""
    parsed = _try_parse_array(value) or []
    return [_defense_reward_struct_dict(item) for item in parsed]


def _analyze_defense_reward_layout(headers, data_rows):
    """统计整张表的最大池数、固定奖励数和随机奖励数。"""
    try:
        score_ci = next(
            ci for ci, field in enumerate(headers)
            if str(field or "").strip() == DEFENSE_REWARD_FIELD
        )
    except StopIteration:
        raise ValueError(f"缺少字段 {DEFENSE_REWARD_FIELD}")

    parsed_rows = []
    max_pool_count = 0
    fixed_counts = {}
    box_counts = {}

    for row in data_rows:
        raw_value = row[score_ci] if score_ci < len(row) else None
        pools = _defense_reward_struct_array(raw_value)
        parsed_pools = []
        max_pool_count = max(max_pool_count, len(pools))

        for pool_idx, pool in enumerate(pools):
            fixed_boxes = _defense_reward_struct_array(pool.get("FixedBoxList"))
            boxes = _defense_reward_struct_array(pool.get("BoxList"))
            fixed_counts[pool_idx] = max(
                fixed_counts.get(pool_idx, 0), len(fixed_boxes)
            )
            box_counts[pool_idx] = max(
                box_counts.get(pool_idx, 0), len(boxes)
            )
            parsed_pools.append({
                "FixedBoxList": fixed_boxes,
                "RandomDropBoxNum": pool.get("RandomDropBoxNum"),
                "BoxList": boxes,
            })
        parsed_rows.append(parsed_pools)

    # 即使当前数组为空，也保留一组可编辑列。
    max_pool_count = max(max_pool_count, 1)
    for pool_idx in range(max_pool_count):
        fixed_counts[pool_idx] = max(fixed_counts.get(pool_idx, 0), 1)
        box_counts[pool_idx] = max(box_counts.get(pool_idx, 0), 1)

    return score_ci, parsed_rows, max_pool_count, fixed_counts, box_counts


def _build_defense_reward_plan(
    headers, score_ci, max_pool_count, fixed_counts, box_counts
):
    """生成 DefenseRewardDefineT_S 专用三行表头和列计划。"""
    row1, row2, row3, plan = [], [], [], []

    for ci, field in enumerate(headers):
        if ci == score_ci:
            for pool_idx in range(max_pool_count):
                parent = f"{DEFENSE_REWARD_FIELD}_{pool_idx}"

                for fixed_idx in range(fixed_counts[pool_idx]):
                    row1.append(parent)
                    row2.append(f"FixedBoxList_GiftBoxTid_{fixed_idx}")
                    row3.append(DEFENSE_REWARD_MARKER)
                    plan.append(("defense_fixed_tid", ci, fixed_idx, pool_idx))

                row1.append(parent)
                row2.append("RandomDropBoxNum")
                row3.append(DEFENSE_REWARD_MARKER)
                plan.append(("defense_random_num", ci, None, pool_idx))

                for box_idx in range(box_counts[pool_idx]):
                    row1.append(parent)
                    row2.append(f"BoxList_GiftBoxTid_{box_idx}")
                    row3.append(DEFENSE_REWARD_MARKER)
                    plan.append(("defense_box_tid", ci, box_idx, pool_idx))

                    row1.append(parent)
                    row2.append(f"BoxList_Weight_{box_idx}")
                    row3.append(DEFENSE_REWARD_MARKER)
                    plan.append(("defense_box_weight", ci, box_idx, pool_idx))
        else:
            row1.append(field)
            row2.append(None)
            row3.append(None)
            plan.append(("normal", ci, None, None))

    row3[0] = "#FieldType"
    return row1, row2, row3, plan


def _expand_defense_reward_row(row, parsed_pools, plan):
    """将一行奖励数据按专用列计划展开。"""
    result = []
    for kind, source_ci, item_idx, pool_idx in plan:
        if kind == "normal":
            result.append(row[source_ci] if source_ci < len(row) else None)
            continue

        pool = parsed_pools[pool_idx] if pool_idx < len(parsed_pools) else None
        if pool is None:
            result.append(None)
            continue

        if kind == "defense_random_num":
            result.append(pool.get("RandomDropBoxNum"))
            continue

        list_name = "FixedBoxList" if kind == "defense_fixed_tid" else "BoxList"
        items = pool.get(list_name) or []
        item = items[item_idx] if item_idx < len(items) else None
        if item is None:
            result.append(None)
        elif kind in ("defense_fixed_tid", "defense_box_tid"):
            result.append(item.get("GiftBoxTid"))
        else:
            result.append(item.get("Weight"))
    return result


def style_header_row(ws, row_num, bg_hex, fg_hex):
    fill  = PatternFill("solid", fgColor=bg_hex)
    font  = Font(color=fg_hex, bold=True)
    align = Alignment(horizontal="center", vertical="center")
    for cell in ws[row_num]:
        cell.fill      = fill
        cell.font      = font
        cell.alignment = align


def auto_col_width(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)


def apply_section_borders(ws, col_plan, total_rows):
    groups = {}
    for new_col_idx, plan in enumerate(col_plan):
        ptype, ci, subkey, arr_idx = plan
        if ptype not in (
            "array_struct", "array_scalar", "tmap_key", "tmap_val", "tset_item",
            "struct_string", "struct_value", "struct_raw",
            "defense_fixed_tid", "defense_random_num",
            "defense_box_tid", "defense_box_weight"
        ):
            continue
        key = (ci, arr_idx)
        if key not in groups:
            groups[key] = []
        groups[key].append(new_col_idx + 1)

    section_starts = set()
    section_ends   = set()
    for key, cols in groups.items():
        section_starts.add(min(cols))
        section_ends.add(max(cols))

    for row in ws.iter_rows(min_row=1, max_row=total_rows,
                             min_col=1, max_col=len(col_plan)):
        for cell in row:
            c     = cell.column
            left  = THICK_SIDE if c in section_starts else NO_SIDE
            right = THICK_SIDE if c in section_ends   else NO_SIDE
            if left.style or right.style:
                eb = cell.border
                cell.border = Border(
                    left   = left  if left.style  else eb.left,
                    right  = right if right.style else eb.right,
                    top    = eb.top,
                    bottom = eb.bottom,
                )

    for cell in ws[total_rows]:
        eb = cell.border
        cell.border = Border(
            left   = eb.left,
            right  = eb.right,
            top    = eb.top,
            bottom = THICK_SIDE,
        )


def process_xlsx(input_path, output_path):
    wb_in    = openpyxl.load_workbook(input_path, read_only=True)
    ws_in    = wb_in.active
    all_rows = list(ws_in.iter_rows(values_only=True))

    field_types = {}
    if TYPE_SHEET_NAME in wb_in.sheetnames:
        ws_meta = wb_in[TYPE_SHEET_NAME]
        for field_name, field_type in ws_meta.iter_rows(min_row=2, values_only=True):
            if field_name and field_type:
                field_types[str(field_name)] = str(field_type).strip().upper()
    wb_in.close()

    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    if not all_rows:
        shutil.copy2(input_path, output_path)
        return "空表，原样复制"

    headers = list(all_rows[0])
    has_type_row = (
        len(all_rows) >= 3 and
        str(all_rows[2][0]).strip() == "#FieldType"
    )
    data_rows = all_rows[3:] if has_type_row else all_rows[1:]
    asset_name = os.path.splitext(os.path.basename(input_path))[0]
    array_cols = analyze_array_columns(headers, data_rows)
    struct_cols = analyze_selected_struct_columns(asset_name, headers, data_rows)

    row1, row2, col_plan = build_column_plan(headers, array_cols, struct_cols)
    row3 = []
    for item in col_plan:
        ptype, ci = item[0], item[1]
        field = str(headers[ci]) if ci < len(headers) and headers[ci] is not None else ""
        if ptype in ("struct_string", "struct_value", "struct_raw"):
            marker = ptype.upper()
        else:
            marker = field_types.get(field)
            if marker is None and ptype != "normal":
                marker = "MAP" if array_cols.get(ci, {}).get("kind") == "tmap" else "ARRAY"
        row3.append(marker)
    row3[0] = "#FieldType"

    wb_out     = openpyxl.Workbook()
    ws_out     = wb_out.active
    ws_out.title = asset_name[:31]

    ws_out.append(row1)
    ws_out.append(row2)
    ws_out.append(row3)
    for row in data_rows:
        ws_out.append(expand_row(row, col_plan))

    style_header_row(ws_out, 1, *STYLE_ROW1)
    style_header_row(ws_out, 2, *STYLE_ROW2)
    style_header_row(ws_out, 3, *STYLE_ROW3)
    auto_col_width(ws_out)
    apply_section_borders(ws_out, col_plan, total_rows=len(data_rows) + 3)

    wb_out.save(output_path)
    array_fields = [d["field"] for d in array_cols.values()]
    struct_fields = [d["field"] for d in struct_cols.values()]
    return (
        f"拆表完成，数组字段: {array_fields}，"
        f"单结构体字段: {struct_fields}，总列数: {len(row1)}"
    )


def process_defense_reward_xlsx(input_path, output_path):
    """
    专用拆分 DefenseRewardDefineT_S。

    ScoreRewardPool 的每个实际数据都会进入独立单元格：
      - FixedBoxList[].GiftBoxTid
      - RandomDropBoxNum
      - BoxList[].GiftBoxTid
      - BoxList[].Weight
    """
    wb_in = openpyxl.load_workbook(input_path, read_only=True)
    ws_in = wb_in.active
    all_rows = list(ws_in.iter_rows(values_only=True))
    wb_in.close()

    if not all_rows:
        raise ValueError("空表")

    headers = list(all_rows[0])
    has_type_row = (
        len(all_rows) >= 3 and
        str(all_rows[2][0]).strip() == "#FieldType"
    )
    data_rows = all_rows[3:] if has_type_row else all_rows[1:]

    score_ci, parsed_rows, max_pool_count, fixed_counts, box_counts = (
        _analyze_defense_reward_layout(headers, data_rows)
    )
    row1, row2, row3, plan = _build_defense_reward_plan(
        headers, score_ci, max_pool_count, fixed_counts, box_counts
    )

    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = DEFENSE_REWARD_TABLE[:31]
    ws_out.append(row1)
    ws_out.append(row2)
    ws_out.append(row3)
    for row, parsed_pools in zip(data_rows, parsed_rows):
        ws_out.append(_expand_defense_reward_row(row, parsed_pools, plan))

    style_header_row(ws_out, 1, *STYLE_ROW1)
    style_header_row(ws_out, 2, *STYLE_ROW2)
    style_header_row(ws_out, 3, *STYLE_ROW3)
    auto_col_width(ws_out)
    apply_section_borders(ws_out, plan, total_rows=len(data_rows) + 3)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb_out.save(output_path)
    return (
        f"专用拆分完成，奖励池 {max_pool_count} 组，"
        f"总列数 {len(row1)}"
    )


def process_csv_rows_directly(asset_name, csv_headers, csv_data, field_types, output_path):
    """
    只供 DIRECT_CSV_SPLIT_TABLES 使用：
    在 UE 原始 CSV 数据仍完整时直接拆列并写入 XLSX，
    避免先写入单个超长 Excel 单元格造成截断。
    """
    headers = ["---"] + list(csv_headers[1:])
    data_rows = []
    for csv_row in csv_data:
        row = list(csv_row)
        if len(row) < len(headers):
            row.extend([None] * (len(headers) - len(row)))
        data_rows.append(row[:len(headers)])

    array_cols = analyze_array_columns(headers, data_rows)
    struct_cols = analyze_selected_struct_columns(asset_name, headers, data_rows)
    row1, row2, col_plan = build_column_plan(headers, array_cols, struct_cols)

    row3 = []
    for item in col_plan:
        ptype, ci = item[0], item[1]
        field = str(headers[ci]) if ci < len(headers) and headers[ci] is not None else ""
        if ptype in ("struct_string", "struct_value", "struct_raw"):
            marker = ptype.upper()
        else:
            marker = field_types.get(field)
            if marker is None and ptype != "normal":
                marker = "MAP" if array_cols.get(ci, {}).get("kind") == "tmap" else "ARRAY"
        row3.append(marker)
    row3[0] = "#FieldType"

    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = asset_name[:31]
    ws_out.append(row1)
    ws_out.append(row2)
    ws_out.append(row3)
    for row in data_rows:
        ws_out.append(expand_row(row, col_plan))

    style_header_row(ws_out, 1, *STYLE_ROW1)
    style_header_row(ws_out, 2, *STYLE_ROW2)
    style_header_row(ws_out, 3, *STYLE_ROW3)
    auto_col_width(ws_out)
    apply_section_borders(ws_out, col_plan, total_rows=len(data_rows) + 3)
    wb_out.save(output_path)

    array_fields = [d["field"] for d in array_cols.values()]
    return f"CSV内存直拆完成，数组字段: {array_fields}，总列数: {len(row1)}"


# ══════════════════════════════════════════════════════
# 【第四部分】拆表主流程
# ══════════════════════════════════════════════════════

def run_split(table_list):
    """拆分本次 UI 勾选并成功导出的表。"""
    if not table_list:
        unreal.log_warning("本次导出选择为空，跳过拆表")
        return 0, []

    unreal.log(f"\n开始拆表，共 {len(table_list)} 个表...")
    success, fail = [], []

    with unreal.ScopedSlowTask(len(table_list), "正在拆表...") as task:
        task.make_dialog(True)

        for entry in table_list:
            task.enter_progress_frame(1, f"拆表中 ({len(success)}/{len(table_list)})：{entry}")

            if task.should_cancel():
                unreal.log_warning("用户取消了拆表操作")
                break

            input_path  = os.path.join(EXPORT_BASE, entry + ".xlsx").replace("\\", "/")
            output_path = os.path.join(CEHUA_BASE,  entry + ".xlsx").replace("\\", "/")

            if not os.path.exists(input_path):
                fail.append(f"{entry}（找不到源文件）")
                unreal.log_warning(f"找不到: {input_path}")
                continue

            try:
                asset_name = os.path.basename(entry)
                if asset_name in DIRECT_CSV_SPLIT_TABLES:
                    os.makedirs(os.path.dirname(output_path), exist_ok=True)
                    shutil.copy2(input_path, output_path)
                    msg = "已使用导出阶段的CSV内存直拆结果"
                elif asset_name == DEFENSE_REWARD_TABLE:
                    msg = process_defense_reward_xlsx(input_path, output_path)
                else:
                    msg = process_xlsx(input_path, output_path)
                success.append(entry)
                unreal.log(f"✅ {entry} - {msg}")
            except Exception as e:
                fail.append(f"{entry}（{e}）")
                unreal.log_error(f"❌ {entry} → {e}")

    unreal.log(f"\n拆表完成：成功 {len(success)} 个，失败 {len(fail)} 个")
    return len(success), fail


# ══════════════════════════════════════════════════════
# 主入口
# ══════════════════════════════════════════════════════

unreal.log("=" * 50)
unreal.log("开始执行：导出并拆表（UI 选表版）")
unreal.log("=" * 50)

try:
    grouped_tables, all_table_entries = scan_exportable_tables()
except Exception as e:
    unreal.log_error(f"扫描 DataTable 失败：{e}")
    unreal.EditorDialog.show_message(
        "扫描失败",
        f"无法读取 UE DataTable 列表：\n\n{e}",
        unreal.AppMsgType.OK,
    )
else:
    if not all_table_entries:
        unreal.log_warning(f"在 {UE_BASE_PATH} 下未找到 DataTable")
        unreal.EditorDialog.show_message(
            "没有可导出的表",
            f"在以下目录未找到 DataTable：\n\n{UE_BASE_PATH}",
            unreal.AppMsgType.OK,
        )
    else:
        valid_entries = set(all_table_entries)
        preselected = load_last_selection() & valid_entries
        selected_tables, open_output_folder = select_export_tables_gui(
            grouped_tables,
            preselected,
        )

        if selected_tables is None:
            unreal.log_warning("用户取消了导出表选择")
        elif not selected_tables:
            unreal.log_warning("未勾选任何表，本次导出结束")
            unreal.EditorDialog.show_message(
                "未选择导出表",
                "本次没有勾选任何 DataTable。",
                unreal.AppMsgType.OK,
            )
        else:
            save_last_selection(selected_tables)
            unreal.log(f"本次 UI 勾选 {len(selected_tables)} 个表")

            export_success, export_fail, exported_list = run_export(selected_tables)
            run_archive(exported_list, catalog_entries=all_table_entries)
            split_success, split_fail = run_split(exported_list)

            msg = (
                f"导出并拆表完成！\n\n"
                f"【本次选择】{len(selected_tables)} 个表\n"
                f"【导出】成功 {export_success} 个，失败 {len(export_fail)} 个\n"
                f"【拆表】成功 {split_success} 个，失败 {len(split_fail)} 个\n\n"
                f"✅ 留档镜像：DataTables_Export_留档\n"
                f"✅ 全量目录：全量表目录.txt\n"
                f"✅ 策划文件：DataTables_Cehua"
            )

            all_fail = []
            if export_fail:
                all_fail.append("导出失败：\n" + "\n".join(export_fail[:5]))
            if split_fail:
                all_fail.append("拆表失败：\n" + "\n".join(split_fail[:5]))
            if all_fail:
                msg += "\n\n⚠️ 失败详情：\n" + "\n\n".join(all_fail)

            unreal.EditorDialog.show_message("导出并拆表完成", msg, unreal.AppMsgType.OK)

            if open_output_folder:
                try:
                    os.startfile(os.path.normpath(CEHUA_BASE))
                except Exception as e:
                    unreal.log_warning(f"打开策划文件夹失败：{e}")
