# -*- coding: utf-8 -*-
"""
导出并拆表 v2.1（UE 5.7 兼容）
流程：
  1. 导出 UE 所有 DataTable → DataTables_Export
  2. 镜像备份              → DataTables_Export_留档
  3. 生成全量表目录.txt
  4. 按操作列表拆表        → DataTables_Cehua
"""

import unreal
import os
import re
import csv as _csv
import json
import shutil
import datetime
import sys

# ── 依赖检测 ──────────────────────────────────────────
def _ensure_openpyxl():
    try:
        import openpyxl
        unreal.log(f"openpyxl 已就绪 (版本 {openpyxl.__version__})")
        return True
    except ImportError:
        unreal.log_error("未检测到 openpyxl")
        unreal.EditorDialog.show_message(
            "缺少依赖：openpyxl",
            "请先运行桌面上的「安装环境.py」脚本安装依赖，\n"
            "安装完成后再运行本脚本。",
            unreal.AppMsgType.OK
        )
        return False

if not _ensure_openpyxl():
    raise SystemExit("缺少依赖 openpyxl，脚本终止")

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles.borders import Border, Side

# ── 路径配置 ──────────────────────────────────────────
BASE_DIR     = os.path.dirname(os.path.abspath(__file__)).replace("\\", "/")
EXPORT_BASE  = f"{BASE_DIR}/DataTables_Export/"
ARCHIVE_BASE = f"{BASE_DIR}/DataTables_Export_留档/"
CEHUA_BASE   = f"{BASE_DIR}/DataTables_Cehua/"
LIST_FILE    = f"{BASE_DIR}/操作列表.txt"
CATALOG_FILE = f"{BASE_DIR}/全量表目录.txt"
UE_BASE_PATH = "/Game/GDataTables"
TYPE_SHEET_NAME = "__UE_META__"

# ── 样式配置 ──────────────────────────────────────────
STYLE_ROW1 = ("2F5496", "FFFFFF")  # 深蓝底 白字
STYLE_ROW2 = ("4472C4", "FFFFFF")  # 中蓝底 白字
STYLE_ROW3 = ("D9E1F2", "000000")  # 浅蓝底 黑字（字段类型）
THICK_SIDE  = Side(style="medium")
NO_SIDE     = Side(style=None)


# ══════════════════════════════════════════════════════
# UE CSV 格式处理工具
# ══════════════════════════════════════════════════════

def _normalize_ue_csv_cell(s):
    """
    把 UE CSV 单元格里的值标准化：
    - UE CSV 数组格式 ((k=v,...),(k=v,...)) 原样保留，
      拆表逻辑的 parse_ue_struct / collect_subkeys 能直接识别
    - 非数组值原样返回
    """
    if not s:
        return s
    s = s.strip()
    return s


def _to_snake_case(name):
    """把 UE 属性名转换为 Python Editor Property 常用的 snake_case。"""
    s1 = re.sub(r'(.)([A-Z][a-z]+)', r'\1_\2', str(name))
    return re.sub(r'([a-z0-9])([A-Z])', r'\1_\2', s1).lower()


def _classify_container_value(value):
    """根据 RowStruct 默认实例中的实际值判断容器类型。"""
    type_name = type(value).__name__.lower()
    if type_name in ("array", "set"):
        return "ARRAY"
    if type_name == "map":
        return "MAP"

    try:
        if isinstance(value, (unreal.Array, unreal.Set)):
            return "ARRAY"
        if isinstance(value, unreal.Map):
            return "MAP"
    except Exception:
        pass
    return None


def get_row_struct_field_types(data_table, export_headers):
    """
    从 DataTable 的 RowStruct 默认实例读取字段真实类型。
    返回 {CSV导出字段名: "ARRAY" | "MAP"}；无法反射的字段留给原有内容分析兜底。
    """
    result = {}
    try:
        row_struct = unreal.DataTableFunctionLibrary.get_data_table_row_struct(data_table)
        if row_struct is None:
            return result

        struct_name = str(row_struct.get_name())
        candidates = [struct_name]
        if struct_name.startswith("F") and len(struct_name) > 1:
            candidates.append(struct_name[1:])

        struct_type = None
        for candidate in candidates:
            struct_type = getattr(unreal, candidate, None)
            if struct_type is not None:
                break
        if struct_type is None:
            unreal.log_warning(f"RowStruct 未生成 Python 类型，改用内容识别: {struct_name}")
            return result

        default_row = struct_type()
        for export_name in export_headers:
            raw_name = unreal.DataTableFunctionLibrary.get_data_table_column_name_from_export_name(
                data_table, str(export_name)
            )
            raw_name = str(raw_name) if raw_name is not None else str(export_name)
            value = None
            found = False
            for prop_name in (raw_name, _to_snake_case(raw_name)):
                try:
                    value = default_row.get_editor_property(prop_name)
                    found = True
                    break
                except Exception:
                    continue
            if found:
                marker = _classify_container_value(value)
                if marker:
                    result[str(export_name)] = marker
    except Exception as e:
        unreal.log_warning(f"读取 RowStruct 字段类型失败，改用内容识别: {e}")
    return result


# ══════════════════════════════════════════════════════
# 【第一部分】导出：UE → DataTables_Export
# ══════════════════════════════════════════════════════

def run_export(table_list):
    # 清空并重建导出文件夹
    if os.path.exists(EXPORT_BASE):
        shutil.rmtree(EXPORT_BASE)
        unreal.log("已清空导出文件夹")
    os.makedirs(EXPORT_BASE, exist_ok=True)

    # 清空策划文件夹
    if os.path.exists(CEHUA_BASE):
        shutil.rmtree(CEHUA_BASE)
        unreal.log("已清空策划文件夹")
    os.makedirs(CEHUA_BASE, exist_ok=True)

    total         = len(table_list)
    success_count = 0
    fail_list     = []
    exported_list = []

    with unreal.ScopedSlowTask(total, "正在导出 DataTable...") as task:
        task.make_dialog(True)

        for entry in table_list:
            asset_name    = os.path.basename(entry)
            relative      = entry
            relative_dir  = os.path.dirname(entry)
            ue_asset_path = f"{UE_BASE_PATH}/{entry}"

            task.enter_progress_frame(1, f"导出中 ({success_count}/{total})：{asset_name}")

            if task.should_cancel():
                unreal.log_warning("用户取消了导出操作")
                break

            output_dir = os.path.join(EXPORT_BASE, relative_dir)
            os.makedirs(output_dir, exist_ok=True)

            try:
                data_table = unreal.load_asset(ue_asset_path)
                if data_table is None:
                    fail_list.append(relative)
                    unreal.log_error(f"找不到UE资产: {ue_asset_path}")
                    continue

                # UE 5.7 兼容：使用 CSV 接口，绕开 JSON 序列化崩溃
                temp_csv = os.path.join(output_dir, f"{asset_name}_temp.csv")
                csv_ok = unreal.DataTableFunctionLibrary.export_data_table_to_csv_file(data_table, temp_csv)
                if not csv_ok or not os.path.exists(temp_csv):
                    fail_list.append(relative)
                    unreal.log_error(f"CSV导出失败: {relative}")
                    continue

                with open(temp_csv, "r", encoding="utf-8") as f:
                    reader = _csv.reader(f)
                    all_csv_rows = list(reader)
                os.remove(temp_csv)

                if not all_csv_rows:
                    fail_list.append(relative)
                    unreal.log_error(f"CSV为空: {relative}")
                    continue

                csv_headers = all_csv_rows[0]   # 第一行：列名
                csv_data    = all_csv_rows[1:]   # 其余：数据行

                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = asset_name[:31]

                # 从 UE RowStruct 读取真实容器类型，写入隐藏元数据表。
                field_types = get_row_struct_field_types(data_table, csv_headers[1:])
                ws_meta = wb.create_sheet(TYPE_SHEET_NAME)
                ws_meta.append(["FieldName", "FieldType"])
                for col_name in csv_headers[1:]:
                    ws_meta.append([col_name, field_types.get(str(col_name), None)])
                ws_meta.sheet_state = "hidden"

                header_fill = PatternFill("solid", fgColor="2F5496")
                header_font = Font(color="FFFFFF", bold=True)

                # 写表头（第一列是 RowName，对应旧版的 "---"）
                ws.cell(row=1, column=1, value="---").fill = header_fill
                ws.cell(row=1, column=1).font = header_font
                for col_idx, col_name in enumerate(csv_headers[1:], start=2):
                    cell = ws.cell(row=1, column=col_idx, value=col_name)
                    cell.fill = header_fill
                    cell.font = header_font

                # 写数据（UE CSV 第一列是 RowName，其余原样写入）
                for row_idx, row in enumerate(csv_data, start=2):
                    for col_idx, val in enumerate(row, start=1):
                        ws.cell(row=row_idx, column=col_idx, value=val if val else None)

                for col in ws.columns:
                    max_len = max((len(str(c.value or "")) for c in col), default=0)
                    ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

                xlsx_path = os.path.join(output_dir, f"{asset_name}.xlsx")
                wb.save(xlsx_path)

                success_count += 1
                exported_list.append(relative)
                unreal.log(f"导出成功: {relative} ({len(csv_data)} 行)")

            except Exception as e:
                fail_list.append(relative)
                unreal.log_error(f"导出失败: {relative} -> {e}")

    unreal.log(f"\n导出完成：成功 {success_count} 个，失败 {len(fail_list)} 个")
    return success_count, fail_list, exported_list


# ══════════════════════════════════════════════════════
# 【第二部分】留档 + 全量表目录
# ══════════════════════════════════════════════════════

def run_archive(exported_list):
    with unreal.ScopedSlowTask(2, "正在生成留档和目录...") as task:
        task.make_dialog(False)

        task.enter_progress_frame(1, "生成留档镜像...")
        if os.path.exists(ARCHIVE_BASE):
            shutil.rmtree(ARCHIVE_BASE)
        shutil.copytree(EXPORT_BASE, ARCHIVE_BASE)
        unreal.log(f"留档镜像已生成: {ARCHIVE_BASE}")

        task.enter_progress_frame(1, "生成全量表目录...")
        dir_groups = {}
        for relative in sorted(exported_list):
            folder = os.path.dirname(relative) or "（根目录）"
            if folder not in dir_groups:
                dir_groups[folder] = []
            dir_groups[folder].append(relative)

        now_str = datetime.datetime.now().strftime("%Y/%m/%d %H:%M")
        lines = [
            f"# 全量表目录 - {now_str}",
            f"# 共 {len(exported_list)} 个表",
            f"# 将需要操作的表名复制到 操作列表.txt 中",
            "",
        ]
        for folder in sorted(dir_groups.keys()):
            lines.append(f"# ── {folder} ──")
            for relative in dir_groups[folder]:
                lines.append(relative)
            lines.append("")

        os.makedirs(os.path.dirname(CATALOG_FILE), exist_ok=True)
        with open(CATALOG_FILE, "w", encoding="utf-8") as f:
            f.write("\n".join(lines))

        unreal.log(f"全量表目录已生成: {CATALOG_FILE} ({len(exported_list)} 个表)")


# ══════════════════════════════════════════════════════
# 【第三部分】拆表核心逻辑
# ══════════════════════════════════════════════════════

def parse_ue_struct(s):
    """解析 (k=v,k=v) 格式，返回 (keys列表, kv字典)（引号内的括号/逗号不参与计数）"""
    s = s.strip()
    if s.startswith("(") and s.endswith(")"):
        s = s[1:-1]
    keys, kv = [], {}
    depth, current = 0, []
    in_q = False
    for ch in s:
        if ch == '"':
            in_q = not in_q
            current.append(ch)
        elif ch in "([" and not in_q:
            depth += 1
            current.append(ch)
        elif ch in ")]" and not in_q:
            depth -= 1
            current.append(ch)
        elif ch == "," and depth == 0 and not in_q:
            token = "".join(current).strip()
            if "=" in token:
                k, _, v = token.partition("=")
                keys.append(k.strip())
                kv[k.strip()] = v.strip().strip('"')
            current = []
        else:
            current.append(ch)
    token = "".join(current).strip()
    if "=" in token:
        k, _, v = token.partition("=")
        keys.append(k.strip())
        kv[k.strip()] = v.strip().strip('"')
    return keys, kv


def collect_subkeys(all_items):
    """从数组元素列表里收集所有子字段名，保持顺序"""
    seen, keys = set(), []
    for item in all_items:
        if isinstance(item, str) and item.strip().startswith("("):
            ks, _ = parse_ue_struct(item)
            for k in ks:
                if k not in seen:
                    seen.add(k)
                    keys.append(k)
    return keys if keys else None


def _split_ue_array_items(inner):
    """把 (k=v,...),(k=v,...) 拆成带括号的元素列表（引号内的括号/逗号不参与计数）"""
    items, depth, current = [], 0, []
    in_q = False
    for ch in inner:
        if ch == '"':
            in_q = not in_q
            current.append(ch)
        elif ch == "(" and not in_q:
            depth += 1
            current.append(ch)
        elif ch == ")" and not in_q:
            depth -= 1
            current.append(ch)
            if depth == 0:
                token = "".join(current).strip()
                if token:
                    items.append(token)
                current = []
        elif ch == "," and depth == 0 and not in_q:
            pass  # 括号外的逗号是元素分隔符，跳过
        else:
            current.append(ch)
    return items


def _parse_kv_pair_item(item_str):
    """
    把 TMap 的 (Key, Value) 格式（括号内无等号，逗号分隔两个值）
    转换为标准 struct 字符串 (Key=key,Value=value)。
    若不符合格式则返回原字符串。
    """
    s = item_str.strip()
    if not (s.startswith("(") and s.endswith(")")):
        return item_str
    inner = s[1:-1]
    # 必须无等号，否则是普通 struct
    if "=" in inner:
        return item_str
    # 按第一个逗号分成两部分（Value 可能含逗号，取首个）
    comma_idx = inner.find(",")
    if comma_idx == -1:
        return item_str
    k = inner[:comma_idx].strip().strip('"')
    v = inner[comma_idx + 1:].strip().strip('"')
    if not k:
        return item_str
    # 重新序列化为标准 (Key=k,Value=v) 格式
    k_part = f'Key="{k}"' if not k.lstrip("-").replace(".", "", 1).isdigit() else f"Key={k}"
    v_part = f'Value="{v}"' if not v.lstrip("-").replace(".", "", 1).isdigit() else f"Value={v}"
    return f"({k_part},{v_part})"


def _try_parse_array(s):
    """
    解析单元格为列表，支持以下格式：
    1. JSON 数组 [...] — 元素可以是 dict / "(k=v)" 字符串 / "(Key,Value)" 字符串 / 标量
    2. UE CSV 结构体数组  ((k=v,...),(k=v,...))  — 双层括号
    3. UE CSV 标量数组    (v1,v2,v3)             — 单层括号，内部无等号且无嵌套括号
    所有结果统一为 "(k=v,...)" 字符串列表（结构体）或标量列表，
    供 collect_subkeys / _detect_col_kind 识别。
    """
    if not s:
        return None
    s = str(s).strip()

    # JSON 格式 [...]
    if s.startswith("[") and s.endswith("]"):
        try:
            parsed = json.loads(s)
            if isinstance(parsed, list):
                result = []
                for item in parsed:
                    if isinstance(item, str) and item.strip().startswith("("):
                        result.append(_parse_kv_pair_item(item.strip()))
                    elif isinstance(item, dict):
                        parts = []
                        for k, v in item.items():
                            if isinstance(v, str):
                                parts.append(f'{k}="{v}"')
                            else:
                                parts.append(f"{k}={v}")
                        result.append(f"({','.join(parts)})")
                    else:
                        result.append(item)
                return result
        except Exception:
            pass

    # UE CSV 格式
    if s.startswith("(") and s.endswith(")"):
        inner = s[1:-1].strip()

        if inner == "":
            # 空数组 ()，返回空列表（不拆列，analyze_array_columns 会跳过空列表）
            return None

        if inner.startswith("("):
            # 双层括号：结构体数组或 TMap
            items = _split_ue_array_items(inner)
            if items:
                return [_parse_kv_pair_item(item) for item in items]
        else:
            # 单层括号：可能是纯标量数组 (v1,v2,v3)
            # 若内部含等号则是单个 struct 值 (k=v,k=v)，不是数组，跳过
            if "=" in inner:
                return None
            items = _split_scalar_items_export(inner)
            if items and len(items) > 1:
                # 多个元素才视为数组，单元素 (v) 视为普通 struct 包装不拆
                return items  # 普通标量列表，_detect_col_kind 会判断为 tset

    return None


def _split_scalar_items_export(inner):
    """
    把 v1,v2,v3 拆成 ['v1','v2','v3']，
    正确处理带括号的嵌套（嵌套括号内的逗号不分割）。
    """
    items, depth, current = [], 0, []
    in_q = False
    for ch in inner:
        if ch == '"':
            in_q = not in_q
            current.append(ch)
        elif ch == "(" and not in_q:
            depth += 1
            current.append(ch)
        elif ch == ")" and not in_q:
            depth -= 1
            current.append(ch)
        elif ch == "," and depth == 0 and not in_q:
            token = "".join(current).strip()
            if token:
                items.append(token)
            current = []
        else:
            current.append(ch)
    token = "".join(current).strip()
    if token:
        items.append(token)
    return items


def _detect_col_kind(all_items):
    """
    检测数组列的种类，返回:
      "tmap"   — 每个元素是 (Key=...,Value=...) struct
      "tset"   — 每个元素是纯标量（数字或字符串），视为 TArray/TSet
      "struct" — 每个元素是其他 struct
      None     — 无法识别或混合格式，不拆列（原样保留）
    规则：
      - 全是 "(k=v,...)" 格式 → 判断子字段决定 tmap / struct
      - 全是非括号标量（int/float/string）→ tset
      - 混合或单元素带括号但无 = → None（单个 struct 值，不是数组）
    """
    if not all_items:
        return None

    struct_items = []
    scalar_items = []
    for item in all_items:
        if isinstance(item, str) and item.strip().startswith("("):
            struct_items.append(item)
        else:
            scalar_items.append(item)

    # 混合格式：不拆
    if struct_items and scalar_items:
        return None

    if scalar_items and not struct_items:
        # 全是非括号元素（int/float/string）→ TArray 标量，拆为 item_N
        return "tset"

    # 全是 struct 元素
    subkeys = collect_subkeys(struct_items)
    if subkeys is None:
        return None
    key_set = set(k.lower() for k in subkeys)
    if key_set == {"key", "value"}:
        return "tmap"
    return "struct"


def analyze_array_columns(headers, data_rows):
    info = {}
    for row in data_rows:
        for ci, val in enumerate(row):
            if val is None:
                continue
            parsed = _try_parse_array(val)
            if parsed is None:
                continue
            if ci not in info:
                info[ci] = {"field": headers[ci], "max_len": 0, "all_items": []}
            if len(parsed) > info[ci]["max_len"]:
                info[ci]["max_len"] = len(parsed)
            info[ci]["all_items"].extend(parsed)
    result = {}
    for ci, d in info.items():
        subkeys = collect_subkeys(d["all_items"])
        kind    = _detect_col_kind(d["all_items"])
        result[ci] = {
            "field":   d["field"],
            "max_len": d["max_len"],
            "subkeys": subkeys,
            "kind":    kind,       # "tmap" | "tset" | "struct"
        }
    return result


def build_column_plan(headers, array_cols):
    """
    生成列计划，plan 元素格式：
      ("normal",       ci, None,    None)
      ("tmap_key",     ci, None,    arr_idx)   — TMap Key
      ("tmap_val",     ci, None,    arr_idx)   — TMap Value
      ("tset_item",    ci, None,    arr_idx)   — TSet 标量元素
      ("array_struct", ci, subkey,  arr_idx)   — 普通结构体数组子字段
      ("array_scalar", ci, None,    arr_idx)   — 普通标量数组
    """
    row1, row2, plan = [], [], []
    for ci, field in enumerate(headers):
        if ci not in array_cols:
            row1.append(field)
            row2.append(None)
            plan.append(("normal", ci, None, None))
            continue

        ac      = array_cols[ci]
        max_len = ac["max_len"]
        subkeys = ac["subkeys"]
        kind    = ac["kind"]
        multi   = max_len > 1

        if kind == "tmap":
            # TMap：每个元素拆成 key_N / val_N 两列
            for arr_idx in range(max_len):
                row1.append(field)
                row2.append(f"key_{arr_idx}" if multi else "key")
                plan.append(("tmap_key", ci, None, arr_idx))
                row1.append(field)
                row2.append(f"val_{arr_idx}" if multi else "val")
                plan.append(("tmap_val", ci, None, arr_idx))

        elif kind == "tset":
            # TSet：每个元素一列，子字段名 item_N
            for arr_idx in range(max_len):
                row1.append(field)
                row2.append(f"item_{arr_idx}" if multi else "item")
                plan.append(("tset_item", ci, None, arr_idx))

        elif kind == "struct" and subkeys:
            # 普通结构体数组
            for arr_idx in range(max_len):
                for sk in subkeys:
                    row1.append(field)
                    row2.append(f"{sk}_{arr_idx}" if multi else sk)
                    plan.append(("array_struct", ci, sk, arr_idx))

        else:
            # 兜底：普通标量数组
            for arr_idx in range(max_len):
                row1.append(field)
                row2.append(f"value_{arr_idx}" if multi else "value")
                plan.append(("array_scalar", ci, None, arr_idx))

    return row1, row2, plan


def _get_tmap_key_value(item):
    """从 TMap 元素中取 Key / Value，兼容大小写"""
    if isinstance(item, str) and item.strip().startswith("("):
        _, kv = parse_ue_struct(item)
        # 找大小写无关的 Key / Value
        key_name = next((k for k in kv if k.lower() == "key"),   None)
        val_name = next((k for k in kv if k.lower() == "value"), None)
        return kv.get(key_name, None), kv.get(val_name, None)
    if isinstance(item, dict):
        key_name = next((k for k in item if k.lower() == "key"),   None)
        val_name = next((k for k in item if k.lower() == "value"), None)
        return item.get(key_name, None), item.get(val_name, None)
    return None, None


def expand_row(row, col_plan):
    new_row      = []
    parsed_cache = {}
    for ptype, ci, subkey, arr_idx in col_plan:
        val = row[ci] if ci < len(row) else None
        if ptype == "normal":
            new_row.append(val)
            continue
        if ci not in parsed_cache:
            parsed_cache[ci] = _try_parse_array(val) or []
        parsed = parsed_cache[ci]
        if arr_idx >= len(parsed):
            new_row.append(None)
            continue
        item = parsed[arr_idx]

        if ptype == "tmap_key":
            k, _ = _get_tmap_key_value(item)
            new_row.append(k)

        elif ptype == "tmap_val":
            _, v = _get_tmap_key_value(item)
            new_row.append(v)

        elif ptype == "tset_item":
            # TSet 标量：直接取元素值
            if isinstance(item, str) and item.strip().startswith("("):
                # 形如 "(123)" 的包装
                inner = item.strip()[1:-1].strip()
                new_row.append(inner if inner else item)
            else:
                new_row.append(item)

        elif ptype == "array_struct":
            if isinstance(item, str) and item.strip().startswith("("):
                _, kv = parse_ue_struct(item)
                new_row.append(kv.get(subkey, None))
            elif isinstance(item, dict):
                v = item.get(subkey, None)
                new_row.append(str(v) if v is not None else None)
            else:
                new_row.append(str(item))

        elif ptype == "array_scalar":
            new_row.append(item)

    return new_row


def style_header_row(ws, row_num, bg_hex, fg_hex):
    fill  = PatternFill("solid", fgColor=bg_hex)
    font  = Font(color=fg_hex, bold=True)
    align = Alignment(horizontal="center", vertical="center")
    for cell in ws[row_num]:
        cell.fill      = fill
        cell.font      = font
        cell.alignment = align


def auto_col_width(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)


def apply_section_borders(ws, col_plan, total_rows):
    groups = {}
    for new_col_idx, plan in enumerate(col_plan):
        ptype, ci, subkey, arr_idx = plan
        if ptype not in ("array_struct", "array_scalar", "tmap_key", "tmap_val", "tset_item"):
            continue
        key = (ci, arr_idx)
        if key not in groups:
            groups[key] = []
        groups[key].append(new_col_idx + 1)

    section_starts = set()
    section_ends   = set()
    for key, cols in groups.items():
        section_starts.add(min(cols))
        section_ends.add(max(cols))

    for row in ws.iter_rows(min_row=1, max_row=total_rows,
                             min_col=1, max_col=len(col_plan)):
        for cell in row:
            c     = cell.column
            left  = THICK_SIDE if c in section_starts else NO_SIDE
            right = THICK_SIDE if c in section_ends   else NO_SIDE
            if left.style or right.style:
                eb = cell.border
                cell.border = Border(
                    left   = left  if left.style  else eb.left,
                    right  = right if right.style else eb.right,
                    top    = eb.top,
                    bottom = eb.bottom,
                )

    for cell in ws[total_rows]:
        eb = cell.border
        cell.border = Border(
            left   = eb.left,
            right  = eb.right,
            top    = eb.top,
            bottom = THICK_SIDE,
        )


def process_xlsx(input_path, output_path):
    wb_in    = openpyxl.load_workbook(input_path, read_only=True)
    ws_in    = wb_in.active
    all_rows = list(ws_in.iter_rows(values_only=True))

    field_types = {}
    if TYPE_SHEET_NAME in wb_in.sheetnames:
        ws_meta = wb_in[TYPE_SHEET_NAME]
        for field_name, field_type in ws_meta.iter_rows(min_row=2, values_only=True):
            if field_name and field_type:
                field_types[str(field_name)] = str(field_type).strip().upper()
    wb_in.close()

    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    if not all_rows:
        shutil.copy2(input_path, output_path)
        return "空表，原样复制"

    headers   = list(all_rows[0])
    data_rows = all_rows[1:]
    array_cols = analyze_array_columns(headers, data_rows)

    row1, row2, col_plan = build_column_plan(headers, array_cols)
    row3 = []
    for item in col_plan:
        ptype, ci = item[0], item[1]
        field = str(headers[ci]) if ci < len(headers) and headers[ci] is not None else ""
        marker = field_types.get(field)
        if marker is None and ptype != "normal":
            marker = "MAP" if array_cols.get(ci, {}).get("kind") == "tmap" else "ARRAY"
        row3.append(marker)
    row3[0] = "#FieldType"

    wb_out     = openpyxl.Workbook()
    ws_out     = wb_out.active
    asset_name = os.path.splitext(os.path.basename(input_path))[0]
    ws_out.title = asset_name[:31]

    ws_out.append(row1)
    ws_out.append(row2)
    ws_out.append(row3)
    for row in data_rows:
        ws_out.append(expand_row(row, col_plan))

    style_header_row(ws_out, 1, *STYLE_ROW1)
    style_header_row(ws_out, 2, *STYLE_ROW2)
    style_header_row(ws_out, 3, *STYLE_ROW3)
    auto_col_width(ws_out)
    apply_section_borders(ws_out, col_plan, total_rows=len(data_rows) + 3)

    wb_out.save(output_path)
    array_fields = [d["field"] for d in array_cols.values()]
    return f"拆表完成，数组字段: {array_fields}，总列数: {len(row1)}"


# ══════════════════════════════════════════════════════
# 【第四部分】拆表主流程
# ══════════════════════════════════════════════════════

def run_split():
    if not os.path.exists(LIST_FILE):
        unreal.log_warning(f"找不到操作列表文件，跳过拆表：{LIST_FILE}")
        unreal.EditorDialog.show_message(
            "未找到操作列表",
            f"导出已完成，但跳过了拆表步骤。\n\n"
            f"请参考桌面的「全量表目录.txt」，\n"
            f"将需要操作的表名复制到：\n\n"
            f"{LIST_FILE}\n\n"
            f"然后重新运行脚本。",
            unreal.AppMsgType.OK
        )
        return 0, []

    with open(LIST_FILE, "r", encoding="utf-8") as f:
        lines = [l.strip() for l in f.readlines()]

    table_list = [
        l.replace("\\", "/").strip()
        for l in lines
        if l and not l.startswith("#")
    ]

    if not table_list:
        unreal.log_warning("操作列表为空，跳过拆表")
        return 0, []

    unreal.log(f"\n开始拆表，共 {len(table_list)} 个表...")
    success, fail = [], []

    with unreal.ScopedSlowTask(len(table_list), "正在拆表...") as task:
        task.make_dialog(True)

        for entry in table_list:
            task.enter_progress_frame(1, f"拆表中 ({len(success)}/{len(table_list)})：{entry}")

            if task.should_cancel():
                unreal.log_warning("用户取消了拆表操作")
                break

            input_path  = os.path.join(EXPORT_BASE, entry + ".xlsx").replace("\\", "/")
            output_path = os.path.join(CEHUA_BASE,  entry + ".xlsx").replace("\\", "/")

            if not os.path.exists(input_path):
                fail.append(f"{entry}（找不到源文件）")
                unreal.log_warning(f"找不到: {input_path}")
                continue

            try:
                msg = process_xlsx(input_path, output_path)
                success.append(entry)
                unreal.log(f"✅ {entry} - {msg}")
            except Exception as e:
                fail.append(f"{entry}（{e}）")
                unreal.log_error(f"❌ {entry} → {e}")

    unreal.log(f"\n拆表完成：成功 {len(success)} 个，失败 {len(fail)} 个")
    return len(success), fail


# ══════════════════════════════════════════════════════
# 主入口
# ══════════════════════════════════════════════════════

unreal.log("=" * 50)
unreal.log("开始执行：导出并拆表")
unreal.log("=" * 50)

if not os.path.exists(LIST_FILE):
    unreal.log("未找到操作列表，执行全量导出...")

    if os.path.exists(EXPORT_BASE):
        shutil.rmtree(EXPORT_BASE)
    os.makedirs(EXPORT_BASE, exist_ok=True)
    if os.path.exists(CEHUA_BASE):
        shutil.rmtree(CEHUA_BASE)
    os.makedirs(CEHUA_BASE, exist_ok=True)

    asset_list = unreal.EditorAssetLibrary.list_assets(UE_BASE_PATH, recursive=True)
    dt_assets  = [
        a for a in asset_list
        if str(unreal.EditorAssetLibrary.find_asset_data(a).asset_class_path.asset_name) == "DataTable"
    ]
    all_entries = []
    for a in dt_assets:
        pkg = str(unreal.EditorAssetLibrary.find_asset_data(a).package_name)
        all_entries.append(pkg.replace(UE_BASE_PATH + "/", ""))

    export_success, export_fail, exported_list = run_export(all_entries)
    run_archive(exported_list)

    unreal.EditorDialog.show_message(
        "全量导出完成",
        f"已全量导出 {export_success} 个表。\n\n"
        f"✅ 留档镜像：DataTables_Export_留档\n"
        f"✅ 全量目录：全量表目录.txt\n\n"
        f"请参考桌面的「全量表目录.txt」，\n"
        f"将需要操作的表名复制到：\n\n"
        f"{LIST_FILE}\n\n"
        f"创建后重新运行脚本，将按操作列表导出并拆表。",
        unreal.AppMsgType.OK
    )

else:
    with open(LIST_FILE, "r", encoding="utf-8") as f:
        lines = [l.strip() for l in f.readlines()]
    table_list = [
        l.replace("\\", "/").strip()
        for l in lines
        if l and not l.startswith("#")
    ]

    if not table_list:
        unreal.log_warning("操作列表为空，退出")
    else:
        unreal.log(f"操作列表共 {len(table_list)} 个表")

        export_success, export_fail, exported_list = run_export(table_list)
        run_archive(exported_list)
        split_success, split_fail = run_split()

        msg = (
            f"导出并拆表完成！\n\n"
            f"【导出】成功 {export_success} 个，失败 {len(export_fail)} 个\n"
            f"【拆表】成功 {split_success} 个，失败 {len(split_fail)} 个\n\n"
            f"✅ 留档镜像：DataTables_Export_留档\n"
            f"✅ 全量目录：全量表目录.txt\n"
            f"✅ 策划文件：DataTables_Cehua"
        )

        all_fail = []
        if export_fail:
            all_fail.append("导出失败：\n" + "\n".join(export_fail[:5]))
        if split_fail:
            all_fail.append("拆表失败：\n" + "\n".join(split_fail[:5]))
        if all_fail:
            msg += "\n\n⚠️ 失败详情：\n" + "\n\n".join(all_fail)

        unreal.EditorDialog.show_message("导出并拆表完成", msg, unreal.AppMsgType.OK)
