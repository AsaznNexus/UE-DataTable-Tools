# -*- coding: utf-8 -*-
"""
合表并导入 v1.2
流程：
  1. 扫描 DataTables_Cehua 下所有表，筛选出在 UE 中有对应资产的【可导入表】
  2. 弹出勾选界面（tkinter），让用户勾选本次要导入的表（默认预勾选 导入列表.txt 中的表）
  3. 按勾选结果合表：DataTables_Cehua → DataTables_Export
  4. P4 冲突检测：找出被他人签出的表
  5. 弹窗确认：有冲突时询问是否跳过继续
  6. 导入：DataTables_Export → UE

说明：
  - 导出/拆表由 操作列表.txt 决定（拆表范围，通常较大），本脚本不依赖
  - 导入列表.txt 仅作为“默认预勾选”的参考，不再是唯一入口
  - 实际导入范围以本次勾选界面的选择为准
"""

import unreal
import os
import re
import sys
import json
import shutil

# ── 依赖检测：openpyxl ────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
except ImportError:
    unreal.log_error("未检测到 openpyxl")
    unreal.EditorDialog.show_message(
        "缺少依赖：openpyxl",
        "请先运行桌面上的「安装环境.py」脚本安装依赖，\n"
        "安装完成后再运行本脚本。",
        unreal.AppMsgType.OK
    )
    raise SystemExit("缺少依赖 openpyxl")

# ── 依赖检测：tkinter ─────────────────────────────────
try:
    import tkinter as tk
except ImportError:
    unreal.log_error("未检测到 tkinter")
    unreal.EditorDialog.show_message(
        "缺少依赖：tkinter",
        "当前 UE 内置 Python 环境未包含 tkinter 模块，\n"
        "无法弹出勾选界面，请联系技术美术检查 Python 环境配置。",
        unreal.AppMsgType.OK
    )
    raise SystemExit("缺少依赖 tkinter")

# ── 路径配置 ──────────────────────────────────────────
BASE_DIR         = os.path.dirname(os.path.abspath(__file__)).replace("\\", "/")
CEHUA_BASE       = f"{BASE_DIR}/DataTables_Cehua/"
EXPORT_BASE      = f"{BASE_DIR}/DataTables_Export/"
LIST_FILE        = f"{BASE_DIR}/操作列表.txt"      # 拆表范围（导出脚本用，本脚本不依赖）
IMPORT_LIST_FILE = f"{BASE_DIR}/导入列表.txt"      # 导入范围（本脚本实际使用）
UE_BASE_PATH = "/Game/GDataTables"
EXCEL_CELL_CHAR_LIMIT = 32767
DIRECT_IMPORT_TABLES = {"DifficultyDefineT"}


class ExcelCellLimitError(Exception):
    """合表后单元格超过 Excel 字符上限时抛出，阻止该表继续导入。"""
    def __init__(self, details):
        self.details = details
        super().__init__(f"有 {len(details)} 个单元格超过 {EXCEL_CELL_CHAR_LIMIT} 字符")


def is_direct_import_table(entry):
    """只对指定特例表启用拆分表直接内存导入。"""
    return os.path.basename(str(entry).replace("\\", "/")) in DIRECT_IMPORT_TABLES


# ══════════════════════════════════════════════════════
# 读取导入列表：扫描全量表 + 勾选界面
# ══════════════════════════════════════════════════════

# ══════════════════════════════════════════════════════
# 读取导入列表：按文件夹分组扫描 + 分组勾选界面
# ══════════════════════════════════════════════════════

def scan_importable_tables(cehua_base, ue_base_path):
    """
    扫描 DataTables_Cehua 下所有 xlsx（支持子文件夹），按【顶层文件夹】分组，
    并筛选出在 UE 中已存在对应 DataTable 资产的表。
    返回 dict：{ 分组名: [entry, entry, ...] }，分组名和组内均按名称排序。
    不在任何子文件夹下的表归入 “（未分类）” 分组。
    """
    groups = {}
    if not os.path.exists(cehua_base):
        return {}

    for root, _, files in os.walk(cehua_base):
        for fn in files:
            if not fn.lower().endswith(".xlsx") or fn.startswith("~$"):
                continue  # 跳过非xlsx及Excel打开时产生的临时锁文件
            full_path = os.path.join(root, fn).replace("\\", "/")
            rel = os.path.relpath(full_path, cehua_base).replace("\\", "/")
            entry = rel[:-5]  # 去掉 .xlsx 后缀

            ue_asset_path = f"{ue_base_path}/{entry}"
            try:
                exists = unreal.EditorAssetLibrary.does_asset_exist(ue_asset_path)
            except Exception as e:
                unreal.log_warning(f"资产存在性查询失败 {entry}：{e}，默认保留候选")
                exists = True
            if not exists:
                continue

            group = entry.split("/", 1)[0] if "/" in entry else "（未分类）"
            groups.setdefault(group, []).append(entry)

    return {g: sorted(es) for g, es in sorted(groups.items())}


def load_default_selection(import_list_file):
    """读取旧版 导入列表.txt（如果存在），作为勾选界面的默认预勾选项。"""
    if not os.path.exists(import_list_file):
        return set()
    with open(import_list_file, "r", encoding="utf-8") as f:
        lines = [l.strip() for l in f.readlines()]
    return {
        l.replace("\\", "/").strip()
        for l in lines
        if l and not l.startswith("#")
    }


def select_tables_gui(grouped, preselected=None):
    """
    弹出 tkinter 勾选窗口，按文件夹分组展示候选表（类似 操作列表.txt 的
    “—— 分组名 ——” 分段风格），支持按组整体勾选/清空、全局搜索过滤。
    grouped: { 分组名: [entry, ...] }
    返回用户确认勾选的 entry 列表（跨所有组）；用户取消/关闭窗口返回 None。
    """
    preselected = preselected or set()
    result = {"selected": None}

    root = tk.Tk()
    root.title("选择要导入的表")
    root.geometry("560x680")
    root.attributes("-topmost", True)

    # ── 顶部：搜索框 + 全局全选/全不选 ──
    top = tk.Frame(root)
    top.pack(fill="x", padx=10, pady=(10, 5))

    tk.Label(top, text="搜索：").pack(side="left")
    search_var = tk.StringVar()
    tk.Entry(top, textvariable=search_var).pack(side="left", fill="x", expand=True, padx=(0, 10))

    count_label = tk.Label(root, text="", anchor="w")
    count_label.pack(fill="x", padx=10)

    # ── 中间：可滚动分组勾选列表 ──
    list_frame = tk.Frame(root)
    list_frame.pack(fill="both", expand=True, padx=10, pady=5)

    canvas    = tk.Canvas(list_frame, highlightthickness=0)
    scrollbar = tk.Scrollbar(list_frame, orient="vertical", command=canvas.yview)
    inner     = tk.Frame(canvas)

    inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
    canvas.create_window((0, 0), window=inner, anchor="nw")
    canvas.configure(yscrollcommand=scrollbar.set)
    canvas.pack(side="left", fill="both", expand=True)
    scrollbar.pack(side="right", fill="y")

    def _on_mousewheel(event):
        canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
    canvas.bind_all("<MouseWheel>", _on_mousewheel)

    vars_by_entry  = {}   # entry -> BooleanVar
    rows_by_entry  = {}   # entry -> Checkbutton widget
    group_headers  = {}   # 分组名 -> 分组标题行 Frame

    def update_count():
        n = sum(v.get() for v in vars_by_entry.values())
        count_label.config(text=f"共 {len(vars_by_entry)} 个可导入表，已选 {n} 个")

    def make_group_setter(group, value):
        def _apply():
            for e in grouped[group]:
                vars_by_entry[e].set(value)
        return _apply

    # ── 按组构建：分组标题 + 组内缩进勾选项 ──
    for group, entries in grouped.items():
        header = tk.Frame(inner, bg="#2F5496")
        header.pack(fill="x", pady=(8, 2))
        tk.Label(header, text=f"—— {group} ——", bg="#2F5496", fg="white",
                 anchor="w").pack(side="left", fill="x", expand=True, padx=6, pady=3)
        tk.Button(header, text="组清空", command=make_group_setter(group, False)
                  ).pack(side="right", padx=(2, 6), pady=2)
        tk.Button(header, text="组全选", command=make_group_setter(group, True)
                  ).pack(side="right", padx=2, pady=2)
        group_headers[group] = header

        for entry in entries:
            var = tk.BooleanVar(value=(entry in preselected))
            var.trace_add("write", lambda *a: update_count())
            # 组内只显示去掉分组前缀的短名，层级感来自缩进
            short_name = entry.split("/", 1)[1] if "/" in entry else entry
            cb = tk.Checkbutton(inner, text=short_name, variable=var, anchor="w", justify="left")
            cb.pack(fill="x", anchor="w", padx=(24, 0))
            vars_by_entry[entry] = var
            rows_by_entry[entry] = cb

    def apply_filter(*_):
        """按关键字过滤；tkinter 的 pack_forget 后重新 pack 会追加到末尾，
        所以这里先把所有条目/标题整体撤下，再按分组顺序依次重新排布，避免顺序错乱。"""
        kw = search_var.get().strip().lower()
        for group, entries in grouped.items():
            group_headers[group].pack_forget()
            for entry in entries:
                rows_by_entry[entry].pack_forget()
        for group, entries in grouped.items():
            visible = [e for e in entries if kw in e.lower()]
            if not visible:
                continue
            group_headers[group].pack(fill="x", pady=(8, 2))
            for entry in visible:
                rows_by_entry[entry].pack(fill="x", anchor="w", padx=(24, 0))
    search_var.trace_add("write", apply_filter)

    def select_all():
        for v in vars_by_entry.values():
            v.set(True)

    def select_none():
        for v in vars_by_entry.values():
            v.set(False)

    tk.Button(top, text="全选",   command=select_all).pack(side="left", padx=2)
    tk.Button(top, text="全不选", command=select_none).pack(side="left", padx=2)

    # ── 底部：确定 / 取消 ──
    bottom = tk.Frame(root)
    bottom.pack(fill="x", padx=10, pady=10)

    def on_confirm():
        result["selected"] = [e for e, v in vars_by_entry.items() if v.get()]
        root.destroy()

    def on_cancel():
        result["selected"] = None
        root.destroy()

    tk.Button(bottom, text="取消", width=10, command=on_cancel).pack(side="right", padx=5)
    tk.Button(bottom, text="确定导入", width=10, command=on_confirm,
              bg="#2F5496", fg="white").pack(side="right", padx=5)

    root.protocol("WM_DELETE_WINDOW", on_cancel)
    update_count()
    root.mainloop()

    return result["selected"]


def load_table_list():
    grouped = scan_importable_tables(CEHUA_BASE, UE_BASE_PATH)
    total = sum(len(es) for es in grouped.values())
    if total == 0:
        msg = (f"在以下目录未找到任何“可导入”的表\n"
               f"（需要同时满足：存在 xlsx 文件 且 UE 中已有对应 DataTable 资产）：\n{CEHUA_BASE}")
        unreal.log_warning(msg)
        unreal.EditorDialog.show_message("无可导入的表", msg, unreal.AppMsgType.OK)
        return []

    preselected = load_default_selection(IMPORT_LIST_FILE)
    selected = select_tables_gui(grouped, preselected)

    if selected is None:
        unreal.log_warning("用户取消了表选择")
        return []
    if not selected:
        unreal.log_warning("未勾选任何表")
    return selected


# ══════════════════════════════════════════════════════
# 【第一部分】合表核心逻辑
# ══════════════════════════════════════════════════════

def parse_subfield_name(s):
    if not s:
        return None
    s = str(s).strip()
    if not s:
        return None
    m = re.match(r'^(.+?)_(\d+)$', s)
    if m:
        return (m.group(1), int(m.group(2)))
    return (s, 0)


def analyze_cehua_headers(row1, row2):
    """
    解析双行表头，返回列计划列表，元素格式：
      ('normal',       ci, field)
      ('tmap_key',     ci, field, arr_idx)   — TMap Key 列
      ('tmap_val',     ci, field, arr_idx)   — TMap Value 列
      ('tset_item',    ci, field, arr_idx)   — TSet 标量元素列
      ('array_scalar', ci, field, arr_idx)   — 普通标量数组列
      ('array_struct', ci, field, arr_idx, subfield) — 普通结构体数组子字段列
    """
    plan = []
    for ci, (f1, f2) in enumerate(zip(row1, row2)):
        f1 = str(f1).strip() if f1 else ""
        f2 = str(f2).strip() if f2 else ""
        if not f2:
            plan.append(('normal', ci, f1))
            continue
        parsed = parse_subfield_name(f2)
        if parsed is None:
            plan.append(('normal', ci, f1))
            continue
        subfield, arr_idx = parsed
        # TMap Key / Value
        if subfield == 'key':
            plan.append(('tmap_key', ci, f1, arr_idx))
        elif subfield == 'val':
            plan.append(('tmap_val', ci, f1, arr_idx))
        # TSet item
        elif subfield == 'item':
            plan.append(('tset_item', ci, f1, arr_idx))
        # 普通标量数组
        elif subfield == 'value':
            plan.append(('array_scalar', ci, f1, arr_idx))
        # 普通结构体数组
        else:
            plan.append(('array_struct', ci, f1, arr_idx, subfield))
    return plan


def get_array_fields(plan):
    seen, fields = set(), []
    for item in plan:
        if item[0] in ('array_struct', 'array_scalar', 'tmap_key', 'tmap_val', 'tset_item'):
            fname = item[2]
            if fname not in seen:
                seen.add(fname)
                fields.append(fname)
    return fields


def _needs_quotes(v):
    """判断 UE CSV struct 写法中该值是否需要加引号"""
    v = str(v) if v is not None else ""
    if v.startswith("(") or v.startswith("{"):
        return False
    try:
        float(v)
        return False
    except ValueError:
        pass
    if v.lower() in ("true", "false", "none"):
        return False
    return True


def _coerce_scalar(v):
    """把单元格值尽量还原为 JSON 兼容标量类型"""
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip()
    if s == "":
        return None
    low = s.lower()
    if low == "true":
        return True
    if low == "false":
        return False
    if low in ("none", "null"):
        return None
    try:
        if "." in s:
            return float(s)
        return int(s)
    except (ValueError, TypeError):
        return s


def restore_row_v2(data_row, plan):
    """
    把策划拆列格式还原为单列值：
    - TMap → JSON 数组 [{"Key":k,"Value":v}, ...]
    - TSet → JSON 数组 [v1, v2, ...]
    - 普通结构体数组 → UE CSV 字符串 "(k=v,...),(k=v,...)" 包裹在 json.dumps 里
    - 普通标量数组 → JSON 数组 [v1, v2, ...]
    """
    field_order = []
    field_plan  = {}
    seen_fields = set()

    for item in plan:
        fname = item[2]
        ftype_key = item[0]
        if ftype_key == 'normal':
            if fname not in seen_fields:
                field_order.append(('normal', fname))
                seen_fields.add(fname)
                field_plan[fname] = [item]
            else:
                field_plan[fname].append(item)
        else:
            if fname not in seen_fields:
                # 记录字段归属种类（tmap/tset/struct/scalar）
                if ftype_key in ('tmap_key', 'tmap_val'):
                    kind = 'tmap'
                elif ftype_key == 'tset_item':
                    kind = 'tset'
                elif ftype_key == 'array_struct':
                    kind = 'struct'
                else:
                    kind = 'scalar'
                field_order.append((kind, fname))
                seen_fields.add(fname)
                field_plan[fname] = []
            field_plan[fname].append(item)

    output_row = []

    for ftype, fname in field_order:
        items = field_plan[fname]

        if ftype == 'normal':
            ci  = items[0][1]
            val = data_row[ci] if ci < len(data_row) else None
            output_row.append(val)

        elif ftype == 'tmap':
            # 按 arr_idx 配对 key/val
            key_map = {}   # arr_idx -> key值
            val_map = {}   # arr_idx -> val值
            for item in items:
                ci  = item[1]
                v   = data_row[ci] if ci < len(data_row) else None
                idx = item[3]
                if item[0] == 'tmap_key':
                    key_map[idx] = _coerce_scalar(v)
                elif item[0] == 'tmap_val':
                    val_map[idx] = _coerce_scalar(v)
            result = []
            for idx in sorted(set(list(key_map.keys()) + list(val_map.keys()))):
                k = key_map.get(idx)
                v = val_map.get(idx)
                # 跳过 key 和 value 均为空的条目
                if k is None and v is None:
                    continue
                result.append({"Key": k, "Value": v})
            # TMap → UE 要求 {k:v,...} Object 格式
            if result:
                tmap_obj = {str(e["Key"]): e["Value"] for e in result}
                output_row.append(json.dumps(tmap_obj, ensure_ascii=False))
            else:
                output_row.append("{}")

        elif ftype == 'tset':
            # TSet：收集 item_N 值
            idx_map = {}
            for item in items:
                ci  = item[1]
                v   = data_row[ci] if ci < len(data_row) else None
                idx = item[3]
                if v is not None and str(v).strip() != "":
                    idx_map[idx] = _coerce_scalar(v)
            result = [idx_map[i] for i in sorted(idx_map.keys())]
            output_row.append(json.dumps(result, ensure_ascii=False) if result else "[]")

        elif ftype == 'struct':
            # 普通结构体数组
            subfield_order = {}
            idx_map        = {}
            for item in items:
                ci       = item[1]
                arr_idx  = item[3]
                subfield = item[4]
                val      = data_row[ci] if ci < len(data_row) else None
                if arr_idx not in subfield_order:
                    subfield_order[arr_idx] = []
                if subfield not in subfield_order[arr_idx]:
                    subfield_order[arr_idx].append(subfield)
                if arr_idx not in idx_map:
                    idx_map[arr_idx] = {}
                idx_map[arr_idx][subfield] = str(val) if val is not None else ""
            # 过滤全空条目
            valid_idx = {i: e for i, e in idx_map.items()
                         if any(v for v in e.values())}
            if not valid_idx:
                output_row.append("[]")
                continue
            result_list = []
            for i in sorted(valid_idx.keys()):
                entry = valid_idx[i]
                parts = []
                for k in subfield_order.get(i, list(entry.keys())):
                    v_str = entry.get(k, "")
                    if _needs_quotes(v_str):
                        v_str = f'"{v_str}"'
                    parts.append(f"{k}={v_str}")
                result_list.append(f"({','.join(parts)})")
            output_row.append(json.dumps(result_list, ensure_ascii=False))

        else:  # scalar
            idx_map = {}
            for item in items:
                ci  = item[1]
                val = data_row[ci] if ci < len(data_row) else None
                idx = item[3]
                if val is not None and str(val).strip() != "":
                    idx_map[idx] = val  # 先保留原始值，稍后判断

            if not idx_map:
                output_row.append("[]")
                continue

            # 检测是否全部是 (Key, Value) 旧版 TMap 格式
            kv_parsed = {}
            all_kv = True
            for idx, raw in idx_map.items():
                parsed_kv = _try_parse_kv_pair(raw)
                if parsed_kv is None:
                    all_kv = False
                    break
                kv_parsed[idx] = parsed_kv

            if all_kv and kv_parsed:
                # 旧版 TMap 格式 → 还原为 UE 要求的 {k:v,...} Object 格式
                tmap_obj = {str(kv_parsed[i]["Key"]): kv_parsed[i]["Value"]
                            for i in sorted(kv_parsed.keys())}
                output_row.append(json.dumps(tmap_obj, ensure_ascii=False))
            else:
                # 普通标量数组
                result = [_coerce_scalar(idx_map[i]) for i in sorted(idx_map.keys())]
                output_row.append(json.dumps(result, ensure_ascii=False))

    return output_row


def _try_parse_kv_pair(v):
    """
    尝试把 (Key, Value) 格式的字符串解析为 {"Key":k,"Value":v}。
    这是 UE 对 TMap 的旧版序列化格式（括号内逗号分隔，无等号）。
    成功返回 dict，失败返回 None。
    """
    if v is None:
        return None
    s = str(v).strip()
    if not (s.startswith("(") and s.endswith(")")):
        return None
    inner = s[1:-1]
    if "=" in inner:
        return None  # 有等号是普通 struct，不是此格式
    comma_idx = inner.find(",")
    if comma_idx == -1:
        return None
    k = inner[:comma_idx].strip().strip('"')
    val_s = inner[comma_idx + 1:].strip().strip('"')
    if not k:
        return None
    # 尝试把 Value 转数字
    coerced_v = _coerce_scalar(val_s)
    return {"Key": k, "Value": coerced_v}


def is_split_table(r1, r2):
    """
    判断是否是拆列格式（第二行有子字段名）。
    支持的子字段名格式：
      普通结构体：  SubFieldName / SubFieldName_N
      TMap：        key / key_N / val / val_N
      TSet：        item / item_N
      普通标量：    value / value_N
    """
    non_none   = [str(v).strip() for v in r2 if v is not None and str(v).strip()]
    none_count = sum(1 for v in r2 if v is None)
    if not non_none or none_count == 0:
        return False
    def is_subfield_name(s):
        return bool(re.match(r'^[A-Za-z][A-Za-z0-9_]*(_\d+)?$', s))
    return all(is_subfield_name(v) for v in non_none)


def _find_oversized_cells(rows, headers):
    """检查合表后将要写入 xlsx 的数据，返回超过 Excel 单元格上限的明细。"""
    details = []
    for row in rows:
        row_name = row[0] if row and row[0] is not None else "<空行名>"
        for ci, value in enumerate(row):
            if value is None:
                continue
            char_count = len(str(value))
            if char_count > EXCEL_CELL_CHAR_LIMIT:
                field_name = headers[ci] if ci < len(headers) else f"第 {ci + 1} 列"
                details.append({
                    "row": str(row_name),
                    "field": str(field_name),
                    "chars": char_count,
                })
    return details


def merge_xlsx(input_path, output_path):
    wb_in    = openpyxl.load_workbook(input_path, read_only=True)
    ws_in    = wb_in.active
    all_rows = list(ws_in.iter_rows(values_only=True))
    wb_in.close()

    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    if not all_rows or len(all_rows) < 3:
        raise ValueError("缺少完整的三行表头")

    row1 = list(all_rows[0])
    row2 = list(all_rows[1])
    row3 = list(all_rows[2])

    if str(row3[0]).strip() != "#FieldType":
        raise ValueError("第三行缺少 #FieldType，数据必须从第四行开始")

    if not is_split_table(row1, row2):
        oversized = _find_oversized_cells(all_rows[3:], row1)
        if oversized:
            raise ExcelCellLimitError(oversized)
        shutil.copy2(input_path, output_path)
        return "无数组字段，原样复制"

    plan         = analyze_cehua_headers(row1, row2)
    array_fields = get_array_fields(plan)

    field_types = {}
    for ci, item in enumerate(plan):
        fname = item[2]
        marker = row3[ci] if ci < len(row3) else None
        marker = str(marker).strip().upper() if marker is not None else ""
        if marker in ("ARRAY", "MAP"):
            field_types[fname] = marker

    if not array_fields:
        oversized = _find_oversized_cells(all_rows[3:], row1)
        if oversized:
            raise ExcelCellLimitError(oversized)
        shutil.copy2(input_path, output_path)
        return "无数组字段，原样复制"

    # 还原原始表头
    original_headers = []
    seen = set()
    for item in plan:
        fname = item[2]
        if fname not in seen:
            seen.add(fname)
            original_headers.append(fname)

    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    asset_name = os.path.splitext(os.path.basename(input_path))[0]
    ws_out.title = asset_name[:31]

    header_fill = PatternFill("solid", fgColor="2F5496")
    header_font = Font(color="FFFFFF", bold=True)
    ws_out.append(original_headers)
    for cell in ws_out[1]:
        cell.fill = header_fill
        cell.font = header_font

    # 合表后仍保留三行表头，导入数据固定从第四行开始。
    ws_out.append([None] * len(original_headers))
    merged_type_row = [field_types.get(name, None) for name in original_headers]
    merged_type_row[0] = "#FieldType"
    ws_out.append(merged_type_row)

    restored_rows = [restore_row_v2(row, plan) for row in all_rows[3:]]
    oversized = _find_oversized_cells(restored_rows, original_headers)
    if oversized:
        raise ExcelCellLimitError(oversized)

    for row in restored_rows:
        ws_out.append(row)

    for col in ws_out.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws_out.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    wb_out.save(output_path)
    return f"合表完成，还原字段: {array_fields}"


# ══════════════════════════════════════════════════════
def validate_direct_import_xlsx(input_path):
    """验证特例表的三行表头，不执行合表。"""
    wb_in = openpyxl.load_workbook(input_path, read_only=True)
    ws_in = wb_in.active
    header_rows = list(ws_in.iter_rows(min_row=1, max_row=3, values_only=True))
    wb_in.close()

    if len(header_rows) < 3:
        raise ValueError("缺少完整的三行表头")
    if not header_rows[2] or str(header_rows[2][0]).strip() != "#FieldType":
        raise ValueError("第三行缺少 #FieldType，数据必须从第四行开始")
    return "三行表头验证通过，将直接从拆分表导入"


def build_direct_import_rows(input_path):
    """
    直接从策划拆分表构建 UE JSON 数据行。
    还原后的超长数组只存在于 Python 内存，不再写入 xlsx 单元格。
    """
    wb_in = openpyxl.load_workbook(input_path, read_only=True)
    ws_in = wb_in.active
    all_rows = list(ws_in.iter_rows(values_only=True))
    wb_in.close()

    if len(all_rows) < 3:
        raise ValueError("缺少完整的三行表头")

    row1 = list(all_rows[0])
    row2 = list(all_rows[1])
    row3 = list(all_rows[2])
    if not row3 or str(row3[0]).strip() != "#FieldType":
        raise ValueError("第三行缺少 #FieldType")

    plan = analyze_cehua_headers(row1, row2)

    original_headers = []
    seen = set()
    for item in plan:
        fname = item[2]
        if fname not in seen:
            seen.add(fname)
            original_headers.append(fname)

    if not original_headers or len(original_headers) < 2:
        raise ValueError("表头为空")

    field_types = {}
    for ci, item in enumerate(plan):
        fname = item[2]
        marker = row3[ci] if ci < len(row3) else None
        marker = str(marker).strip().upper() if marker is not None else ""
        if marker in ("ARRAY", "MAP"):
            field_types[fname] = marker

    array_fields = {name for name, marker in field_types.items() if marker == "ARRAY"}
    map_fields = {name for name, marker in field_types.items() if marker == "MAP"}

    rows = []
    max_cell_chars = 0
    max_cell_row = ""
    max_cell_field = ""

    for data_row in all_rows[3:]:
        restored = restore_row_v2(data_row, plan)
        if not restored:
            continue
        row_name = restored[0]
        if row_name is None:
            continue

        row_data = {"Name": str(row_name)}
        for value_idx, col_name in enumerate(original_headers[1:], start=1):
            value = restored[value_idx] if value_idx < len(restored) else None
            value_chars = len(str(value)) if value is not None else 0
            if value_chars > max_cell_chars:
                max_cell_chars = value_chars
                max_cell_row = str(row_name)
                max_cell_field = str(col_name)
            row_data[col_name] = parse_cell_value(
                value,
                force_array=(col_name in array_fields),
                force_map=(col_name in map_fields)
            )
        rows.append(row_data)

    if not rows:
        raise ValueError("无数据行")

    return rows, {
        "max_cell_chars": max_cell_chars,
        "max_cell_row": max_cell_row,
        "max_cell_field": max_cell_field,
    }


# 【第二部分】P4 冲突检测
# ══════════════════════════════════════════════════════

def check_p4_conflicts(table_list):
    can_import = []
    conflicts  = []
    source_control = unreal.SourceControlHelpers

    for entry in table_list:
        ue_asset_path = f"{UE_BASE_PATH}/{entry}"
        try:
            state = source_control.query_file_state(ue_asset_path)
            if state is None:
                can_import.append(entry)
                continue
            if state.is_checked_out_other:
                other_user = getattr(state, 'other_user_checked_out', '其他人')
                conflicts.append((entry, str(other_user)))
            else:
                can_import.append(entry)
        except Exception as e:
            unreal.log_warning(f"P4状态查询失败 {entry}: {e}，默认允许导入")
            can_import.append(entry)

    return can_import, conflicts


# ══════════════════════════════════════════════════════
# 【第三部分】导入核心逻辑
# ══════════════════════════════════════════════════════

def _split_ue_array_items(inner):
    """把 (k=v,...),(k=v,...) 拆成带括号的元素列表（引号内的括号/逗号不参与计数）"""
    items, depth, current = [], 0, []
    in_q = False
    for ch in inner:
        if ch == '"':
            in_q = not in_q
            current.append(ch)
        elif ch == "(" and not in_q:
            depth += 1
            current.append(ch)
        elif ch == ")" and not in_q:
            depth -= 1
            current.append(ch)
            if depth == 0:
                token = "".join(current).strip()
                if token:
                    items.append(token)
                current = []
        elif ch == "," and depth == 0 and not in_q:
            pass  # 括号外的逗号是元素分隔符，跳过
        else:
            current.append(ch)
    return items


def _ue_struct_to_dict(s):
    """把 (k=v,k=v) 解析成 dict，值尝试转数字"""
    s = s.strip().strip("()")
    if not s:
        return None
    pairs, current, in_q, depth = [], [], False, 0
    for ch in s:
        if ch == '"':
            in_q = not in_q
            current.append(ch)
        elif ch == "(" and not in_q:
            depth += 1
            current.append(ch)
        elif ch == ")" and not in_q:
            depth -= 1
            current.append(ch)
        elif ch == "," and not in_q and depth == 0:
            pairs.append("".join(current).strip())
            current = []
        else:
            current.append(ch)
    if current:
        pairs.append("".join(current).strip())
    result = {}
    for pair in pairs:
        if "=" not in pair:
            continue
        k, _, v = pair.partition("=")
        k = k.strip()
        v = v.strip().strip('"')
        try:
            v = int(v)
        except ValueError:
            try:
                v = float(v)
            except ValueError:
                pass
        result[k] = v
    return result if result else None


def _is_tmap_list(obj):
    """
    判断一个 Python list 是否是 TMap 的 [{Key:..., Value:...}] 格式。
    """
    if not isinstance(obj, list) or not obj:
        return False
    return all(
        isinstance(item, dict) and
        len(item) == 2 and
        any(k.lower() == "key"   for k in item) and
        any(k.lower() == "value" for k in item)
        for item in obj
    )


def _tmap_list_to_obj(lst):
    """
    把 [{Key:k, Value:v}, ...] 转成 UE fill_data_table 要求的 {k: v, ...} 对象。
    """
    result = {}
    for item in lst:
        key_name = next(k for k in item if k.lower() == "key")
        val_name = next(k for k in item if k.lower() == "value")
        result[str(item[key_name])] = item[val_name]
    return result


def _is_single_ue_struct(inner):
    """
    判断括号内是否是【单个 UE 结构体】的 key=value 序列，
    例如 AttributeName="X",Attribute=...,AttributeOwner="..."
    依据：按顶层逗号拆分后，存在形如  标识符=...  的 token。
    （顶层 = 不在更深一层括号内；_split_scalar_items 已正确处理嵌套）
    标量集合 (1,2,3) / ("a","b") 不含顶层 = → 返回 False。
    """
    for token in _split_scalar_items(inner):
        if re.match(r'^[A-Za-z_]\w*\s*=', token.strip()):
            return True
    return False


def parse_cell_value(val, force_array=False, force_map=False):
    """
    把 xlsx 单元格值转成适合 fill_data_table_from_json_string 的 Python 对象。
    支持：
    - TMap JSON [{Key:k,Value:v},...] → {k:v,...}  (UE 要求 Object 格式)
    - JSON 格式 [...] / {...}         → 直接解析
    - UE TMap CSV ((Key=k,Value=v),...)  → {k:v,...}
    - UE TMap CSV ((Key,Value) 无等号)   → {k:v,...}
    - UE TSet CSV (v1,v2,...)            → [v1,v2,...]
    - UE 普通结构体数组 ((k=v,...),...)   → [{"k":v,...},...]
    - 普通数字/字符串

    force_array：由 import_table 的整列扫描决定。某列若任意一行出现过 [...] 或 ((
    （数组特征），则该列判定为【数组字段】，force_array=True。
    用于消解「单层括号 (键=值,...)」的二义性：
      - 数组字段（如 SkillIds）：裸 (k=v,...) 是单元素，应包成数组 ["(k=v,...)"]；
        空括号 () 应还原为空数组 []。否则触发 "Expected Array, got String"。
      - 单结构体字段（如 GameplayAttribute）：force_array=False，裸 (k=v,...) 保持 String。
    """
    # 空单元格需要根据字段类型返回正确的空值。
    # UE 的 TArray 字段要求 []；普通字段继续使用空字符串。
    if val is None:
        if force_map:
            return {}
        return [] if force_array else ""

    s = str(val).strip()
    if s == "":
        if force_map:
            return {}
        return [] if force_array else ""

    # JSON 格式
    if (s.startswith("[") and s.endswith("]")) or \
       (s.startswith("{") and s.endswith("}")):
        try:
            parsed = json.loads(s)
            # [{Key:k,Value:v},...] → TMap → 转成 {k:v,...}
            if _is_tmap_list(parsed):
                return _tmap_list_to_obj(parsed)
            return parsed
        except Exception:
            pass

    # UE CSV 数组格式 ((...),...) 或 (v1,v2,...) 或空括号 ()
    if s.startswith("(") and s.endswith(")"):
        inner = s[1:-1].strip()

        # 空括号 () → 空集合/空 map。
        # 数组字段还原为空数组 []；其余保持空 dict（UE 对空 TMap 可接受 {}）。
        if inner == "":
            return [] if force_array else {}

        # 单个结构体 (Key=Val,Key=Val,...)：括号内是顶层 key=value 序列，
        # 既不是结构体数组 ((...),(...))，也不是标量集合 (v1,v2,...)。
        #   - 单结构体字段（GameplayAttribute）：原样返回 String，交给 UE ImportText；
        #     若误拆成数组，会触发 "Expected String, got Array"。
        #   - 数组字段（SkillIds 等）单元素场景：force_array=True，包成 ["(k=v,...)"]
        #     单元素数组（与合表输出格式一致），否则触发 "Expected Array, got String"。
        if not inner.startswith("(") and _is_single_ue_struct(inner):
            return [s] if force_array else s

        if inner.startswith("("):
            # 结构体数组：((Key=k,Value=v),...) 或 ((Key, Value) 无等号格式,...) 或 ((field=v,...)...)
            raw_items = _split_ue_array_items(inner)
            if raw_items:
                result = []
                is_tmap = None
                for item in raw_items:
                    # 先尝试 (Key, Value) 无等号格式
                    kv = _try_parse_kv_pair(item)
                    if kv is not None:
                        if is_tmap is None:
                            is_tmap = True
                        result.append(kv)
                        continue
                    d = _ue_struct_to_dict(item)
                    if d is None:
                        result.append(item)
                        is_tmap = False
                        continue
                    # 检测是否 TMap（Key + Value 两个字段）
                    if is_tmap is None:
                        lkeys = set(k.lower() for k in d.keys())
                        is_tmap = (lkeys == {"key", "value"})
                    if is_tmap:
                        key_name = next((k for k in d if k.lower() == "key"),   None)
                        val_name = next((k for k in d if k.lower() == "value"), None)
                        result.append({
                            "Key":   d.get(key_name),
                            "Value": d.get(val_name),
                        })
                    else:
                        result.append(d)
                if result:
                    # TMap → 转成 UE 要求的 {k:v,...} Object 格式
                    if is_tmap:
                        return _tmap_list_to_obj(result)
                    return result

        else:
            # 纯标量列表：(v1,v2,...) → TSet
            parts = _split_scalar_items(inner)
            if parts:
                result = []
                for p in parts:
                    p = p.strip().strip('"')
                    try:
                        result.append(int(p))
                    except ValueError:
                        try:
                            result.append(float(p))
                        except ValueError:
                            result.append(p)
                return result

    # 数字
    try:
        if "." in s:
            return float(s)
        return int(s)
    except Exception:
        pass

    return s


def _split_scalar_items(inner):
    """把 v1,v2,v3 拆成 ['v1','v2','v3']，正确处理带括号的嵌套"""
    items, depth, current = [], 0, []
    for ch in inner:
        if ch == "(":
            depth += 1
            current.append(ch)
        elif ch == ")":
            depth -= 1
            current.append(ch)
        elif ch == "," and depth == 0:
            token = "".join(current).strip()
            if token:
                items.append(token)
            current = []
        else:
            current.append(ch)
    token = "".join(current).strip()
    if token:
        items.append(token)
    return items


def import_table(entry):
    """导入单张表，返回 (success, message)"""
    xlsx_path     = os.path.join(EXPORT_BASE, entry + ".xlsx")
    ue_asset_path = f"{UE_BASE_PATH}/{entry}"

    if not os.path.exists(xlsx_path):
        return False, "找不到xlsx文件"

    try:
        data_table = unreal.load_asset(ue_asset_path)
        if data_table is None or not isinstance(data_table, unreal.DataTable):
            return False, "找不到UE资产"

        wb      = openpyxl.load_workbook(xlsx_path)
        ws      = wb.active
        headers = []

        for col in range(2, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            if val:
                headers.append(str(val))

        if not headers:
            return False, "表头为空"

        # 所有文件统一使用三行表头，数据固定从第四行开始。
        if str(ws.cell(row=3, column=1).value).strip() != "#FieldType":
            return False, "第三行缺少 #FieldType"
        data_start = 4
        array_fields = set()
        map_fields = set()

        for col_idx, col_name in enumerate(headers, start=2):
            marker = ws.cell(row=3, column=col_idx).value
            marker = str(marker).strip().upper() if marker is not None else ""
            if marker == "ARRAY":
                array_fields.add(col_name)
            elif marker == "MAP":
                map_fields.add(col_name)

        rows = []
        for row in range(data_start, ws.max_row + 1):
            row_name = ws.cell(row=row, column=1).value
            if row_name is None:
                continue
            row_data = {"Name": str(row_name)}
            for col_idx, col_name in enumerate(headers, start=2):
                val = ws.cell(row=row, column=col_idx).value
                row_data[col_name] = parse_cell_value(
                    val,
                    force_array=(col_name in array_fields),
                    force_map=(col_name in map_fields)
                )
            rows.append(row_data)

        if not rows:
            return False, "无数据行"

        json_str = json.dumps(rows, ensure_ascii=False)
        result   = unreal.DataTableFunctionLibrary.fill_data_table_from_json_string(
            data_table, json_str
        )

        if result:
            try:
                unreal.SourceControlHelpers.check_out_file(ue_asset_path)
            except Exception:
                pass
            save_ok = unreal.EditorAssetLibrary.save_asset(ue_asset_path)
            if save_ok:
                return True, f"{len(rows)} 行"
            return False, "fill_data_table 成功，但资产保存失败"
        else:
            return False, "fill_data_table 失败"

    except Exception as e:
        return False, str(e)


# ══════════════════════════════════════════════════════
def import_direct_table(entry):
    """特例表直接从 DataTables_Cehua 内存还原并导入 UE。"""
    xlsx_path = os.path.join(CEHUA_BASE, entry + ".xlsx")
    ue_asset_path = f"{UE_BASE_PATH}/{entry}"

    if not os.path.exists(xlsx_path):
        return False, "找不到策划xlsx文件"

    try:
        data_table = unreal.load_asset(ue_asset_path)
        if data_table is None or not isinstance(data_table, unreal.DataTable):
            return False, "找不到UE资产"

        rows, direct_stats = build_direct_import_rows(xlsx_path)
        json_str = json.dumps(rows, ensure_ascii=False)
        result = unreal.DataTableFunctionLibrary.fill_data_table_from_json_string(
            data_table, json_str
        )

        if result:
            try:
                unreal.SourceControlHelpers.check_out_file(ue_asset_path)
            except Exception:
                pass
            save_ok = unreal.EditorAssetLibrary.save_asset(ue_asset_path)
            if save_ok:
                return True, (
                    f"直接内存导入 {len(rows)} 行，JSON {len(json_str)} 字符，"
                    f"最长还原字段 {direct_stats['max_cell_chars']} 字符"
                )
            return False, "fill_data_table 成功，但资产保存失败"
        return False, "fill_data_table 失败"

    except Exception as e:
        return False, str(e)


def format_limit_skips(limit_skipped, max_items=10):
    """格式化超过 Excel 单元格字符上限的跳过明细。"""
    lines = []
    total_cells = sum(len(details) for _, details in limit_skipped)
    for entry, details in limit_skipped:
        for detail in details:
            if len(lines) >= max_items:
                break
            lines.append(
                f"  • {entry} | 行 {detail['row']} | 字段 {detail['field']} | "
                f"{detail['chars']} 字符"
            )
        if len(lines) >= max_items:
            break
    if total_cells > len(lines):
        lines.append(f"  …其余 {total_cells - len(lines)} 个超限单元格请查看 Output Log")
    return "\n".join(lines), total_cells


# 主流程
# ══════════════════════════════════════════════════════

unreal.log("=" * 50)
unreal.log("开始执行：合表并导入")
unreal.log("=" * 50)

def run_preflight_check():
    issues = []
    if not os.path.exists(CEHUA_BASE):
        issues.append(f"❌ 找不到策划文件夹：\n   {CEHUA_BASE}\n   请先运行「导出并拆表」脚本")
    if not os.path.exists(EXPORT_BASE):
        issues.append(f"❌ 找不到导出文件夹：\n   {EXPORT_BASE}\n   请先运行「导出并拆表」脚本")
    if issues:
        msg = "前置条件不满足，无法执行：\n\n" + "\n\n".join(issues)
        unreal.log_error(msg)
        unreal.EditorDialog.show_message("前置检查失败", msg, unreal.AppMsgType.OK)
        return False
    unreal.log("✅ 前置检查通过")
    return True

if not run_preflight_check():
    unreal.log("脚本终止")
else:
    table_list = load_table_list()
    if not table_list:
        unreal.log_warning("导入列表为空，退出")
    else:
        total = len(table_list)
        unreal.log(f"导入列表共 {total} 个表")

        merge_success, merge_fail   = [], []
        limit_skipped               = []
        direct_import_ready         = set()
        direct_import_success       = []
        direct_import_fail          = []
        import_success, import_fail = [], []

        # ── 第一步：合表 ──
        with unreal.ScopedSlowTask(total, "正在合表...") as task:
            task.make_dialog(True)
            for entry in table_list:
                task.enter_progress_frame(1, f"合表中 ({len(merge_success)}/{total})：{entry}")
                if task.should_cancel():
                    unreal.log_warning("用户取消了合表操作")
                    break

                input_path  = os.path.join(CEHUA_BASE,  entry + ".xlsx").replace("\\", "/")
                output_path = os.path.join(EXPORT_BASE, entry + ".xlsx").replace("\\", "/")

                if not os.path.exists(input_path):
                    merge_fail.append(f"{entry}（找不到策划文件）")
                    unreal.log_warning(f"找不到: {input_path}")
                    continue

                try:
                    if is_direct_import_table(entry):
                        msg = validate_direct_import_xlsx(input_path)
                        direct_import_ready.add(entry)
                    else:
                        msg = merge_xlsx(input_path, output_path)
                    merge_success.append(entry)
                    if entry in direct_import_ready:
                        unreal.log(f"✅ 直接导入准备: {entry} - {msg}")
                    else:
                        unreal.log(f"✅ 合表: {entry} - {msg}")
                except ExcelCellLimitError as e:
                    limit_skipped.append((entry, e.details))
                    for detail in e.details:
                        unreal.log_warning(
                            f"⚠️ 超限跳过: {entry} | 行 {detail['row']} | "
                            f"字段 {detail['field']} | {detail['chars']} 字符"
                        )
                except Exception as e:
                    merge_fail.append(f"{entry}（{e}）")
                    unreal.log_error(f"❌ 合表失败: {entry} → {e}")

        unreal.log(
            f"\n预处理完成：常规合表 {len(merge_success) - len(direct_import_ready)} 个，"
            f"直接导入准备 {len(direct_import_ready)} 个，"
            f"失败 {len(merge_fail)} 个，超限跳过 {len(limit_skipped)} 个"
        )

        if not merge_success:
            no_success_msg = "没有可继续导入的表。\n请检查 DataTables_Cehua 文件夹。"
            if limit_skipped:
                limit_text, limit_cell_count = format_limit_skips(limit_skipped)
                no_success_msg += (
                    f"\n\n【超过 Excel {EXCEL_CELL_CHAR_LIMIT} 字符限制】"
                    f"\n已跳过 {len(limit_skipped)} 个表，共 {limit_cell_count} 个超限单元格："
                    f"\n{limit_text}"
                )
            unreal.EditorDialog.show_message(
                "没有可导入表",
                no_success_msg,
                unreal.AppMsgType.OK
            )
        else:
            # ── 第二步：P4 冲突检测 ──
            unreal.log("\n正在检测 P4 签出状态...")
            with unreal.ScopedSlowTask(len(merge_success), "检测 P4 冲突...") as task:
                task.make_dialog(False)
                can_import = []
                conflicts  = []
                for entry in merge_success:
                    task.enter_progress_frame(1, f"检测：{entry}")
                    ue_asset_path = f"{UE_BASE_PATH}/{entry}"
                    try:
                        state = unreal.SourceControlHelpers.query_file_state(ue_asset_path)
                        if state and state.is_checked_out_other:
                            other = getattr(state, 'other_user_checked_out', '其他人')
                            conflicts.append((entry, str(other)))
                        else:
                            can_import.append(entry)
                    except Exception as e:
                        unreal.log_warning(f"P4检测失败 {entry}: {e}，默认允许导入")
                        can_import.append(entry)

            # ── 第三步：冲突弹窗 ──
            if conflicts:
                conflict_lines = "\n".join([f"  • {e}（签出人：{u}）" for e, u in conflicts])
                msg = (
                    f"以下 {len(conflicts)} 个表已被他人签出，无法导入：\n\n"
                    f"{conflict_lines}\n\n"
                    f"是否跳过冲突表，继续导入其余 {len(can_import)} 个表？"
                )
                choice = unreal.EditorDialog.show_message(
                    "P4 签出冲突", msg, unreal.AppMsgType.YesNo
                )
                if choice != unreal.AppReturnType.Yes:
                    unreal.log("用户选择中止，导入取消")
                    can_import = []
                else:
                    for entry, user in conflicts:
                        import_fail.append(f"{entry}（已被 {user} 签出）")
                    unreal.log(f"跳过 {len(conflicts)} 个冲突表，继续导入 {len(can_import)} 个")

            # ── 第四步：导入 ──
            if can_import:
                with unreal.ScopedSlowTask(len(can_import), "正在导入到 UE...") as task:
                    task.make_dialog(True)
                    for entry in can_import:
                        task.enter_progress_frame(
                            1, f"导入中 ({len(import_success)}/{len(can_import)})：{entry}"
                        )
                        if task.should_cancel():
                            unreal.log_warning("用户取消了导入操作")
                            break

                        if entry in direct_import_ready:
                            ok, msg = import_direct_table(entry)
                        else:
                            ok, msg = import_table(entry)
                        if ok:
                            import_success.append(entry)
                            if entry in direct_import_ready:
                                direct_import_success.append(entry)
                            unreal.log(f"✅ 导入: {entry} ({msg})")
                        else:
                            import_fail.append(f"{entry}（{msg}）")
                            if entry in direct_import_ready:
                                direct_import_fail.append(entry)
                            unreal.log_error(f"❌ 导入失败: {entry} → {msg}")

            # ── 最终弹窗 ──
            summary = (
                f"合表并导入完成！\n\n"
                f"【常规合表】成功 {len(merge_success) - len(direct_import_ready)} 个，"
                f"失败 {len(merge_fail)} 个\n"
                f"【直接内存导入】成功 {len(direct_import_success)} 个，"
                f"失败 {len(direct_import_fail)} 个\n"
                f"【导入】成功 {len(import_success)} 个，失败 {len(import_fail)} 个\n"
                f"【超限跳过】{len(limit_skipped)} 个表"
            )
            all_fail = []
            if merge_fail:
                all_fail.append("合表失败：\n" + "\n".join(merge_fail[:5]))
            if import_fail:
                all_fail.append("导入失败/跳过：\n" + "\n".join(import_fail[:5]))
            if limit_skipped:
                limit_text, limit_cell_count = format_limit_skips(limit_skipped)
                all_fail.append(
                    f"超过 Excel {EXCEL_CELL_CHAR_LIMIT} 字符限制：\n"
                    f"已跳过 {len(limit_skipped)} 个表，共 {limit_cell_count} 个超限单元格\n"
                    f"{limit_text}"
                )
            if all_fail:
                summary += "\n\n⚠️ 详情：\n" + "\n\n".join(all_fail)

            unreal.log(summary)
            unreal.EditorDialog.show_message("合表并导入完成", summary, unreal.AppMsgType.OK)
